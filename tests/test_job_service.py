from __future__ import annotations

import asyncio
import hashlib
import time
from datetime import UTC, datetime, timedelta
from pathlib import Path
from urllib.parse import parse_qs, urlparse

import pytest
from openpyxl import Workbook

from douhot_crawler import job_service
from douhot_crawler.job_service import JobManager, JobStore, ServiceSettings, user_key


def settings(tmp_path: Path) -> ServiceSettings:
    return ServiceSettings(
        data_root=tmp_path / "data",
        public_url="http://crawler.test",
        download_secret="test-secret",
        login_timeout_seconds=5,
    )


def test_user_key_is_stable_and_rejects_unsafe_values():
    assert user_key("alice") == hashlib.sha256(b"alice").hexdigest()
    assert user_key("alice") != user_key("bob")
    with pytest.raises(ValueError):
        user_key("  ")
    with pytest.raises(ValueError):
        user_key("bad\nuser")


def test_paths_copies_configured_cookie_and_overwrites_existing_user_cookie(tmp_path):
    cookie_source = tmp_path / "cookie.config"
    cookie_source.write_text("initial-cookie", encoding="utf-8")
    service_settings = settings(tmp_path)
    service_settings = ServiceSettings(
        data_root=service_settings.data_root,
        public_url=service_settings.public_url,
        download_secret=service_settings.download_secret,
        login_timeout_seconds=service_settings.login_timeout_seconds,
        cookie_source=cookie_source,
    )
    manager = JobManager(service_settings)

    paths = manager.paths("alice")

    assert paths.cookie.read_text(encoding="utf-8") == "initial-cookie"
    assert paths.cookie.stat().st_mode & 0o777 == 0o600

    paths.cookie.write_text("user-cookie", encoding="utf-8")
    cookie_source.write_text("updated-template", encoding="utf-8")

    refreshed_paths = manager.paths("alice")

    assert refreshed_paths.cookie.read_text(encoding="utf-8") == "updated-template"
    assert refreshed_paths.cookie.stat().st_mode & 0o777 == 0o600


def test_store_isolates_jobs_and_recovers_interrupted_jobs(tmp_path):
    path = tmp_path / "jobs.sqlite3"
    store = JobStore(path)
    alice = user_key("alice")
    bob = user_key("bob")
    job = store.create(owner=alice, kind="crawl", params={"keyword": "AI"})
    store.update(job["id"], alice, status="running")

    with pytest.raises(ValueError, match="任务不存在"):
        store.get(job["id"], bob)
    with pytest.raises(ValueError, match="任务不存在"):
        store.get(job["id"][:8], bob)

    assert store.get(job["id"][:8], alice)["id"] == job["id"]

    recovered = JobStore(path).get(job["id"], alice)
    assert recovered["status"] == "failed"
    assert "服务重启" in recovered["error"]


def test_only_one_active_job_per_user(tmp_path):
    store = JobStore(tmp_path / "jobs.sqlite3")
    owner = user_key("alice")
    first = store.create(owner=owner, kind="crawl", params={})
    with pytest.raises(ValueError, match=first["id"]):
        store.create(owner=owner, kind="login", params={})
    store.update(first["id"], owner, status="succeeded")
    assert store.create(owner=owner, kind="login", params={})["kind"] == "login"


def test_store_latest_job_is_scoped_by_user_and_kind(tmp_path):
    store = JobStore(tmp_path / "jobs.sqlite3")
    alice = user_key("alice")
    bob = user_key("bob")
    old_crawl = store.create(owner=alice, kind="crawl", params={})
    store.update(old_crawl["id"], alice, status="failed")
    login = store.create(owner=alice, kind="login", params={})
    store.update(login["id"], alice, status="succeeded")
    latest_crawl = store.create(owner=alice, kind="crawl", params={})
    store.update(latest_crawl["id"], alice, status="succeeded")
    bob_crawl = store.create(owner=bob, kind="crawl", params={})
    store.update(bob_crawl["id"], bob, status="succeeded")

    assert store.latest(alice, kinds=("crawl",))["id"] == latest_crawl["id"]
    assert store.latest(alice, kinds=("login",))["id"] == login["id"]
    assert store.latest(bob, kinds=("crawl",))["id"] == bob_crawl["id"]


def test_describe_without_id_returns_current_users_latest_crawl_or_analyze(tmp_path):
    manager = JobManager(settings(tmp_path))
    owner = user_key("alice")
    login = manager.store.create(owner=owner, kind="login", params={})
    manager.store.update(login["id"], owner, status="succeeded")
    crawl = manager.store.create(owner=owner, kind="crawl", params={})
    manager.store.update(crawl["id"], owner, status="failed")

    assert manager.describe("alice", None, kinds=("crawl", "analyze"))["id"] == crawl["id"]
    with pytest.raises(ValueError, match="任务不存在"):
        manager.describe("bob", None, kinds=("crawl", "analyze"))


async def test_start_crawl_rejects_unknown_filter_values(tmp_path):
    manager = JobManager(settings(tmp_path))

    with pytest.raises(ValueError, match="result_type"):
        await manager.start_crawl(
            "alice",
            keyword="AI+创业",
            result_type="视频",
            time_range="近7天",
            input_timeout=30,
            detail_delay=1,
        )

    assert manager.store.active(user_key("alice")) is None


async def test_start_crawl_initializes_browser_without_health_call(
    tmp_path, monkeypatch
):
    manager = JobManager(settings(tmp_path))
    checked = False

    def unavailable_browser():
        nonlocal checked
        checked = True
        return False, "尚未安装浏览器"

    monkeypatch.setattr(job_service, "chromium_status", unavailable_browser)

    with pytest.raises(ValueError, match="尚未安装浏览器"):
        await manager.start_crawl(
            "alice",
            keyword="AI+创业",
            result_type="低粉爆款",
            time_range="近7天",
            input_timeout=30,
            detail_delay=1,
        )

    assert checked is True
    assert manager.store.active(user_key("alice")) is None


async def test_successful_empty_crawl_still_creates_excel(tmp_path, monkeypatch):
    manager = JobManager(settings(tmp_path))

    async def empty_crawl(_options, **_kwargs):
        return {"added_count": 0, "skipped_count": 0, "sheet": "AI+创业"}

    monkeypatch.setattr(job_service, "chromium_status", lambda: (True, "ready"))
    monkeypatch.setattr(job_service, "run_crawl", empty_crawl)
    started = await manager.start_crawl(
        "alice",
        keyword="AI+创业",
        result_type="低粉爆款",
        time_range="近7天",
        input_timeout=30,
        detail_delay=1,
    )
    for _ in range(20):
        result = manager.describe("alice", started["id"])
        if result["status"] in {"succeeded", "failed"}:
            break
        await asyncio.sleep(0.01)

    assert result["status"] == "succeeded", result["error"]
    assert result["result"]["row_count"] == 0
    assert Path(result["result_path"]).is_file()


async def test_start_crawl_passes_result_limit_to_crawler(tmp_path, monkeypatch):
    manager = JobManager(settings(tmp_path))
    captured = {}

    async def empty_crawl(_options, **kwargs):
        captured.update(kwargs)
        return {"added_count": 0, "skipped_count": 0, "sheet": "AI+创业"}

    monkeypatch.setattr(job_service, "chromium_status", lambda: (True, "ready"))
    monkeypatch.setattr(job_service, "run_crawl", empty_crawl)
    started = await manager.start_crawl(
        "alice",
        keyword="AI+创业",
        result_type="低粉爆款",
        time_range="近7天",
        input_timeout=30,
        detail_delay=1,
        limit=7,
    )

    for _ in range(20):
        result = manager.describe("alice", started["id"])
        if result["status"] in {"succeeded", "failed"}:
            break
        await asyncio.sleep(0.01)

    assert result["status"] == "succeeded", result["error"]
    assert result["params"]["limit"] == 7
    assert captured["max_results"] == 7


async def test_start_crawl_rejects_non_positive_result_limit(tmp_path):
    manager = JobManager(settings(tmp_path))

    with pytest.raises(ValueError, match="limit"):
        await manager.start_crawl(
            "alice",
            keyword="AI+创业",
            result_type="低粉爆款",
            time_range="近7天",
            input_timeout=30,
            detail_delay=1,
            limit=0,
        )


async def test_shutdown_cancels_and_waits_for_active_jobs(tmp_path):
    manager = JobManager(settings(tmp_path))
    runner_started = asyncio.Event()

    async def runner(_job_id, _owner, cancellation):
        runner_started.set()
        await cancellation.wait()
        return {}

    job = await manager._start(
        user_id="alice",
        kind="crawl",
        params={},
        runner=runner,
    )
    await runner_started.wait()

    await manager.shutdown()

    assert manager.describe("alice", job["id"])["status"] == "cancelled"
    assert manager._tasks == {}


async def test_wait_for_terminal_notifies_until_background_job_completes(tmp_path):
    manager = JobManager(settings(tmp_path))
    release = asyncio.Event()
    statuses = []

    async def runner(_job_id, _owner, _cancellation):
        await release.wait()
        return {"count": 3}

    job = await manager._start(
        user_id="alice",
        kind="crawl",
        params={},
        runner=runner,
    )

    async def on_status(current):
        statuses.append(current["status"])
        release.set()

    result = await manager.wait_for_terminal(
        "alice",
        job["id"],
        timeout=1,
        poll_interval=0.01,
        on_status=on_status,
    )

    assert statuses == ["queued", "succeeded"]
    assert result["status"] == "succeeded"
    assert result["result"] == {"count": 3}


async def test_wait_for_terminal_timeout_does_not_cancel_background_job(tmp_path):
    manager = JobManager(settings(tmp_path))
    release = asyncio.Event()

    async def runner(_job_id, _owner, _cancellation):
        await release.wait()
        return {}

    job = await manager._start(
        user_id="alice",
        kind="crawl",
        params={},
        runner=runner,
    )

    with pytest.raises(TimeoutError, match="仍在后台运行"):
        await manager.wait_for_terminal(
            "alice", job["id"], timeout=0.01, poll_interval=0.005
        )

    assert manager.describe("alice", job["id"])["status"] == "running"
    release.set()
    await manager.wait_for_terminal("alice", job["id"], timeout=1)


async def test_capture_qr_waits_for_square_image_instead_of_screenshotting_page(
    tmp_path,
):
    class Candidate:
        async def is_visible(self):
            return True

        async def bounding_box(self):
            return {"width": 180, "height": 180}

        async def screenshot(self, *, path):
            Path(path).write_bytes(b"qr-image")

    class Candidates:
        def __init__(self, page, selector):
            self.page = page
            self.selector = selector

        async def count(self):
            return int(self.page.ready and self.selector == "img")

        def nth(self, _index):
            return Candidate()

    class Page:
        ready = False
        full_page_screenshot_called = False

        def locator(self, selector):
            return Candidates(self, selector)

        async def wait_for_timeout(self, _milliseconds):
            self.ready = True

        async def screenshot(self, **_kwargs):
            self.full_page_screenshot_called = True

    page = Page()
    output = tmp_path / "qr.png"

    await JobManager._capture_qr(page, output, timeout_ms=1_000)

    assert output.read_bytes() == b"qr-image"
    assert page.full_page_screenshot_called is False


async def test_start_login_reuses_active_login_job(tmp_path):
    manager = JobManager(settings(tmp_path))
    owner = user_key("alice")
    active = manager.store.create(owner=owner, kind="login", params={})
    qr_path = manager.paths("alice").jobs / active["id"] / "qr.png"
    qr_path.parent.mkdir(parents=True)
    qr_path.write_bytes(b"qr-image")
    manager.store.update(
        active["id"],
        owner,
        status="waiting_login",
        result_path=str(qr_path),
        result_mime="image/png",
    )

    result = await manager.start_login("alice")

    assert result["id"] == active["id"]
    assert result["status"] == "waiting_login"
    assert result["download_url"].startswith("http://crawler.test/downloads/")


def test_signed_download_is_scoped_and_expires(tmp_path, monkeypatch):
    manager = JobManager(settings(tmp_path))
    owner = user_key("alice")
    job = manager.store.create(owner=owner, kind="crawl", params={})
    output = manager.paths("alice").jobs / job["id"] / "result.xlsx"
    output.parent.mkdir(parents=True)
    output.write_bytes(b"xlsx")
    manager.store.update(
        job["id"],
        owner,
        status="succeeded",
        result_path=str(output),
        result_mime="application/test",
    )
    url = manager.signed_download_url("alice", job["id"])
    query = parse_qs(urlparse(url).query)
    path, mime = manager.resolve_download(
        job["id"], query["owner"][0], query["expires"][0], query["signature"][0]
    )
    assert path == output.resolve()
    assert mime == "application/test"

    monkeypatch.setattr(time, "time", lambda: int(query["expires"][0]) + 1)
    with pytest.raises(ValueError, match="过期"):
        manager.resolve_download(
            job["id"], query["owner"][0], query["expires"][0], query["signature"][0]
        )


def test_list_videos_returns_stable_selection_fields(tmp_path):
    manager = JobManager(settings(tmp_path))
    owner = user_key("alice")
    job = manager.store.create(owner=owner, kind="crawl", params={})
    output = manager.paths("alice").jobs / job["id"] / "result.xlsx"
    output.parent.mkdir(parents=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["视频名称", "视频的url", "博主名称", "热度值"])
    sheet.append(["标题", "https://www.douyin.com/video/123456", "作者", "99w"])
    workbook.save(output)
    manager.store.update(
        job["id"],
        owner,
        status="succeeded",
        result_path=str(output),
        result_mime="application/test",
    )

    result = manager.list_videos("alice", job["id"], offset=0, limit=20)
    assert result["total"] == 1
    assert result["videos"] == [
        {
            "index": 1,
            "video_id": "123456",
            "title": "标题",
            "author": "作者",
            "heat": "99w",
            "url": "https://www.douyin.com/video/123456",
            "sheet": "Sheet",
        }
    ]


def test_list_videos_finds_latest_cached_crawl_by_keyword(tmp_path):
    manager = JobManager(settings(tmp_path))
    owner = user_key("alice")

    def cached_job(keyword: str, video_id: str):
        job = manager.store.create(
            owner=owner, kind="crawl", params={"keyword": keyword}
        )
        output = manager.paths("alice").jobs / job["id"] / "result.xlsx"
        output.parent.mkdir(parents=True)
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["视频名称", "视频的url", "博主名称", "热度值"])
        sheet.append(
            ["标题", f"https://www.douyin.com/video/{video_id}", "作者", "99w"]
        )
        workbook.save(output)
        manager.store.update(
            job["id"],
            owner,
            status="succeeded",
            result_path=str(output),
            result_mime="application/test",
        )
        return job

    cached_job("健康", "111")
    expected = cached_job("AI 创业", "222")

    result = manager.list_videos(
        "alice", None, keyword=" ai 创业 ", offset=0, limit=20
    )
    missing = manager.list_videos(
        "alice", None, keyword="数据库没有的词", offset=0, limit=20
    )

    assert result["found"] is True
    assert result["job_id"] == expected["id"]
    assert result["keyword"] == "AI 创业"
    assert result["videos"][0]["video_id"] == "222"
    assert missing == {
        "found": False,
        "keyword": "数据库没有的词",
        "total": 0,
        "offset": 0,
        "limit": 20,
        "videos": [],
    }


def test_keyword_video_cache_is_isolated_by_user(tmp_path):
    manager = JobManager(settings(tmp_path))
    owner = user_key("alice")
    job = manager.store.create(
        owner=owner, kind="crawl", params={"keyword": "私有关键词"}
    )
    manager.store.update(job["id"], owner, status="failed")

    result = manager.list_videos(
        "bob", None, keyword="私有关键词", offset=0, limit=20
    )

    assert result["found"] is False


def test_keyword_video_cache_expires_after_four_hours(tmp_path, monkeypatch):
    manager = JobManager(settings(tmp_path))
    owner = user_key("alice")
    expired_at = (datetime.now(UTC) - timedelta(hours=4, seconds=1)).isoformat()

    with monkeypatch.context() as patch:
        patch.setattr(job_service, "utc_now", lambda: expired_at)
        job = manager.store.create(
            owner=owner, kind="crawl", params={"keyword": "大健康"}
        )
        output = manager.paths("alice").jobs / job["id"] / "result.xlsx"
        output.parent.mkdir(parents=True)
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["视频名称", "视频的url", "博主名称", "热度值"])
        sheet.append(
            ["标题", "https://www.douyin.com/video/123456", "作者", "99w"]
        )
        workbook.save(output)
        manager.store.update(
            job["id"],
            owner,
            status="succeeded",
            result_path=str(output),
            result_mime="application/test",
        )

    result = manager.list_videos(
        "alice", None, keyword="大健康", offset=0, limit=20
    )

    assert result == {
        "found": False,
        "keyword": "大健康",
        "total": 0,
        "offset": 0,
        "limit": 20,
        "videos": [],
    }
