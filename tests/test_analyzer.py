from __future__ import annotations

import argparse
import json
from io import BytesIO
from urllib.error import HTTPError

import pytest
from openpyxl import Workbook, load_workbook

from douhot_crawler.transcript import analyzer


def test_extract_video_returns_changed_api_response_fields(monkeypatch) -> None:
    response = {
        "success": True,
        "video_id": "7661611117199106176",
        "title": "真的笑麻了",
        "video_url": "https://aweme.snssdk.com/aweme/v1/play/?video_id=test",
        "transcript": "测试口播",
    }
    monkeypatch.setattr(
        analyzer,
        "urlopen",
        lambda *_, **__: BytesIO(json.dumps(response).encode("utf-8")),
    )

    result = analyzer.extract_video(
        "https://www.douyin.com/video/7661611117199106176",
        "sessionid=test-only",
        "",
        1.0,
        api_url="https://example.test/extract",
    )

    assert result == analyzer.ExtractedVideo(
        video_id="7661611117199106176",
        title="真的笑麻了",
        video_url="https://aweme.snssdk.com/aweme/v1/play/?video_id=test",
        transcript="测试口播",
    )


def test_extract_video_accepts_missing_play_url(monkeypatch) -> None:
    response = {
        "success": True,
        "video_id": "7661611117199106176",
        "title": "真的笑麻了",
        "transcript": "测试口播",
    }
    monkeypatch.setattr(
        analyzer,
        "urlopen",
        lambda *_, **__: BytesIO(json.dumps(response).encode("utf-8")),
    )

    result = analyzer.extract_video(
        "https://www.douyin.com/video/7661611117199106176",
        "sessionid=test-only",
        "",
        1.0,
        api_url="https://example.test/extract",
    )

    assert result.video_url == ""
    assert result.transcript == "测试口播"


def test_extract_transcript_includes_safe_http_error_detail(monkeypatch) -> None:
    error = HTTPError(
        "https://example.test/extract",
        400,
        "bad request",
        {},
        BytesIO(b'{"detail":"cookie sessionid=secret is invalid"}'),
    )
    def raise_error(*args, **kwargs):
        raise error

    monkeypatch.setattr(analyzer, "urlopen", raise_error)

    with pytest.raises(RuntimeError) as caught:
        analyzer.extract_transcript(
            "https://www.douyin.com/video/1",
            "sessionid=secret",
            "",
            1.0,
            api_url="https://example.test/extract",
        )

    assert "HTTP 400" in str(caught.value)
    assert "[REDACTED]" in str(caught.value)
    assert "sessionid=secret" not in str(caught.value)


def test_analyze_logs_running_and_final_success_counts(tmp_path, capsys) -> None:
    excel_path = tmp_path / "result.xlsx"
    cookie_path = tmp_path / "cookie.config"
    cookie_path.write_text("sessionid=test-only", encoding="utf-8")

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "老板创业"
    worksheet.append(["视频的url"])
    worksheet.append(["https://www.douyin.com/video/1"])
    worksheet.append(["https://www.douyin.com/video/2"])
    workbook.save(excel_path)
    workbook.close()

    result = analyzer.analyze_excel(
        argparse.Namespace(
            excel=excel_path,
            cookie_file=cookie_path,
            sheet=None,
            callback_url="",
            timeout=1.0,
            delay=0.0,
            limit=None,
            overwrite=False,
            extractor=lambda **_: "测试口播",
        )
    )

    assert result == (2, 0, 0)
    output = capsys.readouterr().out
    assert "第 2 行口播已写入（本次已完成 1 个）" in output
    assert "第 3 行口播已写入（本次已完成 2 个）" in output
    assert "口播提取完成：成功 2 个，跳过 0 个，失败 0 个" in output


def test_analyze_writes_transcript_and_video_play_url(tmp_path) -> None:
    excel_path = tmp_path / "result.xlsx"
    cookie_path = tmp_path / "cookie.config"
    cookie_path.write_text("sessionid=test-only", encoding="utf-8")

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "AI创业"
    worksheet.append(["视频的url"])
    worksheet.append(["https://www.douyin.com/video/7661611117199106176"])
    workbook.save(excel_path)
    workbook.close()

    result = analyzer.analyze_excel(
        argparse.Namespace(
            excel=excel_path,
            cookie_file=cookie_path,
            sheet=None,
            callback_url="",
            timeout=1.0,
            delay=0.0,
            limit=None,
            overwrite=False,
            extractor=lambda **_: analyzer.ExtractedVideo(
                video_id="7661611117199106176",
                title="真的笑麻了",
                video_url="https://aweme.snssdk.com/aweme/v1/play/?video_id=test",
                transcript="测试口播",
            ),
        )
    )

    workbook = load_workbook(excel_path)
    worksheet = workbook["AI创业"]
    headers = [cell.value for cell in worksheet[1]]
    transcript_cell = worksheet.cell(2, headers.index("视频口播") + 1)
    play_url_cell = worksheet.cell(2, headers.index("视频播放地址") + 1)
    observed = (
        result,
        transcript_cell.value,
        play_url_cell.value,
        play_url_cell.hyperlink.target if play_url_cell.hyperlink else None,
    )
    workbook.close()

    assert observed == (
        (1, 0, 0),
        "测试口播",
        "https://aweme.snssdk.com/aweme/v1/play/?video_id=test",
        "https://aweme.snssdk.com/aweme/v1/play/?video_id=test",
    )


def test_analyze_backfills_play_url_without_overwriting_existing_transcript(
    tmp_path,
) -> None:
    excel_path = tmp_path / "result.xlsx"
    cookie_path = tmp_path / "cookie.config"
    cookie_path.write_text("sessionid=test-only", encoding="utf-8")

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "旧数据"
    worksheet.append(["视频的url", "视频口播"])
    worksheet.append(["https://www.douyin.com/video/1", "原有口播"])
    workbook.save(excel_path)
    workbook.close()

    result = analyzer.analyze_excel(
        argparse.Namespace(
            excel=excel_path,
            cookie_file=cookie_path,
            sheet=None,
            callback_url="",
            timeout=1.0,
            delay=0.0,
            limit=None,
            overwrite=False,
            extractor=lambda **_: analyzer.ExtractedVideo(
                video_id="1",
                title="标题",
                video_url="https://aweme.snssdk.com/aweme/v1/play/?video_id=1",
                transcript="接口返回的新口播",
            ),
        )
    )

    workbook = load_workbook(excel_path)
    worksheet = workbook["旧数据"]
    headers = [cell.value for cell in worksheet[1]]
    observed = (
        result,
        worksheet.cell(2, headers.index("视频口播") + 1).value,
        worksheet.cell(2, headers.index("视频播放地址") + 1).value,
    )
    workbook.close()

    assert observed == (
        (1, 0, 0),
        "原有口播",
        "https://aweme.snssdk.com/aweme/v1/play/?video_id=1",
    )


def test_analyze_success_limit_counts_only_nonblank_new_transcripts(tmp_path) -> None:
    excel_path = tmp_path / "result.xlsx"
    cookie_path = tmp_path / "cookie.config"
    cookie_path.write_text("sessionid=test-only", encoding="utf-8")

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "候选视频"
    worksheet.append(["视频的url"])
    for index in range(1, 5):
        worksheet.append([f"https://www.douyin.com/video/{index}"])
    workbook.save(excel_path)
    workbook.close()

    outcomes = iter(["", "口播 2", "口播 3", "不应调用"])
    calls: list[str] = []

    def extractor(**kwargs):
        calls.append(kwargs["share_link"])
        return next(outcomes)

    result = analyzer.analyze_excel(
        argparse.Namespace(
            excel=excel_path,
            cookie_file=cookie_path,
            sheet=None,
            callback_url="",
            timeout=1.0,
            delay=0.0,
            limit=None,
            success_limit=2,
            overwrite=False,
            extractor=extractor,
        )
    )

    assert result == (2, 0, 1)
    assert calls == [
        "https://www.douyin.com/video/1",
        "https://www.douyin.com/video/2",
        "https://www.douyin.com/video/3",
    ]


def test_analyze_success_limit_skips_existing_transcript_without_backfill(
    tmp_path,
) -> None:
    excel_path = tmp_path / "result.xlsx"
    cookie_path = tmp_path / "cookie.config"
    cookie_path.write_text("sessionid=test-only", encoding="utf-8")

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "恢复任务"
    worksheet.append(["视频的url", "视频口播"])
    worksheet.append(["https://www.douyin.com/video/1", "已有口播"])
    worksheet.append(["https://www.douyin.com/video/2", ""])
    workbook.save(excel_path)
    workbook.close()

    calls: list[str] = []

    def extractor(**kwargs):
        calls.append(kwargs["share_link"])
        return "新增口播"

    result = analyzer.analyze_excel(
        argparse.Namespace(
            excel=excel_path,
            cookie_file=cookie_path,
            sheet=None,
            callback_url="",
            timeout=1.0,
            delay=0.0,
            limit=None,
            success_limit=1,
            overwrite=False,
            extractor=extractor,
        )
    )

    assert result == (1, 1, 0)
    assert calls == ["https://www.douyin.com/video/2"]


def test_analyze_stops_immediately_when_extraction_service_is_unavailable(
    tmp_path,
) -> None:
    excel_path = tmp_path / "result.xlsx"
    cookie_path = tmp_path / "cookie.config"
    cookie_path.write_text("sessionid=test-only", encoding="utf-8")
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["视频的url"])
    worksheet.append(["https://www.douyin.com/video/1"])
    worksheet.append(["https://www.douyin.com/video/2"])
    workbook.save(excel_path)
    workbook.close()
    calls = 0

    def extractor(**kwargs):
        nonlocal calls
        calls += 1
        raise analyzer.ExtractionServiceUnavailable("提取服务不可用")

    with pytest.raises(analyzer.ExtractionServiceUnavailable):
        analyzer.analyze_excel(
            argparse.Namespace(
                excel=excel_path,
                cookie_file=cookie_path,
                sheet=None,
                callback_url="",
                timeout=1.0,
                delay=0.0,
                limit=None,
                success_limit=2,
                overwrite=False,
                extractor=extractor,
            )
        )

    assert calls == 1
