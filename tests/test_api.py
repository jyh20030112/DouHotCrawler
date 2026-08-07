from __future__ import annotations

import asyncio
import sqlite3
import time
from datetime import datetime
from pathlib import Path
from typing import Any
from zoneinfo import ZoneInfo

import httpx
import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook

from douhot_crawler.api.app import create_app
from douhot_crawler.api.clients import ExternalApiClient
from douhot_crawler.api.config import ApiSettings
from douhot_crawler.api import daily
from douhot_crawler.api.models import (
    AnalyzeTaskRequest,
    PipelineDataSource,
    PipelineTaskRequest,
    TaskKind,
    TaskStatus,
    UploadTaskRequest,
)
from douhot_crawler.api.service import ApiTaskService, parse_follower_count
from douhot_crawler.api.store import ApiTaskStore
from douhot_crawler.api.errors import ApiError, ExternalServiceError, TaskPaused
from douhot_crawler.core.config import RESULT_HEADERS
from douhot_crawler.core.storage import atomic_workbook_save, excel_sheet_name


def settings(tmp_path: Path) -> ApiSettings:
    return ApiSettings(
        hotspot_api_url="https://example.test/hotspots",
        industry_api_url="https://example.test/industry",
        cookie_api_url="https://example.test/cookies",
        ranking_api_url="https://example.test/rankings",
        industry_ranking_api_url="https://example.test/industry-rankings",
        extract_api_url="https://example.test/extract",
        hotspot_open_id="test-open-id",
        data_root=tmp_path,
    )


def test_settings_accepts_single_worker_from_environment(
    tmp_path: Path, monkeypatch
) -> None:
    monkeypatch.setenv("DOUHOT_HOTSPOT_API_URL", "https://example.test/hotspots")
    monkeypatch.setenv("DOUHOT_INDUSTRY_API_URL", "https://example.test/industry")
    monkeypatch.setenv("DOUHOT_COOKIE_API_URL", "https://example.test/cookies")
    monkeypatch.setenv("DOUHOT_RANKING_API_URL", "https://example.test/rankings")
    monkeypatch.setenv(
        "DOUHOT_INDUSTRY_RANKING_API_URL",
        "https://example.test/industry-rankings",
    )
    monkeypatch.setenv("EXTRACT_API_URL", "https://example.test/extract")
    monkeypatch.setenv("DOUHOT_HOTSPOT_OPEN_ID", "test-open-id")
    monkeypatch.setenv("DOUHOT_API_DATA_ROOT", str(tmp_path))
    monkeypatch.setenv("DOUHOT_API_WORKERS", "1")
    monkeypatch.setenv("DOUHOT_MAX_VIDEOS_PER_KEYWORD", "17")
    monkeypatch.setenv("DOUHOT_MAX_CANDIDATES_PER_KEYWORD", "31")
    monkeypatch.setenv("DOUHOT_DAILY_ENABLED", "false")
    monkeypatch.setenv("DOUHOT_DAILY_TIME", "04:30")

    configured = ApiSettings()
    assert configured.workers == 1
    assert configured.max_videos_per_keyword == 17
    assert configured.max_candidates_per_keyword == 31
    assert configured.daily_enabled is False
    assert configured.daily_time == "04:30"


def test_pipeline_request_defaults_to_all_data_sources() -> None:
    request = PipelineTaskRequest()

    assert request.data_source == PipelineDataSource.ALL
    assert request.limit_per_keyword is None
    assert request.candidate_limit_per_keyword is None


def test_pipeline_request_rejects_candidate_limit_below_target() -> None:
    with pytest.raises(ValueError, match="候选视频上限不能小于有效口播目标"):
        PipelineTaskRequest(
            data_source="hotspot",
            limit_per_keyword=4,
            candidate_limit_per_keyword=3,
        )


def test_pipeline_rejects_effective_candidate_limit_below_target(
    tmp_path: Path, monkeypatch
) -> None:
    monkeypatch.setattr(
        "douhot_crawler.api.service.chromium_status", lambda: (True, "test browser")
    )
    configured = settings(tmp_path)
    configured.max_videos_per_keyword = 4
    configured.max_candidates_per_keyword = 3
    service = ApiTaskService(configured, client=FakeExternalClient())

    with pytest.raises(ApiError) as caught:
        service.create_pipeline(PipelineTaskRequest(data_source="hotspot"))

    assert caught.value.code == "INVALID_PIPELINE_LIMITS"
    assert caught.value.details == {
        "limit_per_keyword": 4,
        "candidate_limit_per_keyword": 3,
    }


@pytest.mark.asyncio
async def test_analyze_pauses_when_transcript_cookie_service_is_unavailable(
    tmp_path: Path, monkeypatch
) -> None:
    monkeypatch.setattr(
        "douhot_crawler.api.service.chromium_status", lambda: (True, "test browser")
    )

    class CookieFailureClient(FakeExternalClient):
        async def fetch_cookie(self, cookie_type: int) -> str:
            raise ExternalServiceError("Cookie 配置接口", "网络请求失败")

    service = ApiTaskService(settings(tmp_path), client=CookieFailureClient())
    task = service.store.create(TaskKind.ANALYZE, {})

    with pytest.raises(TaskPaused, match="transcript_cookie_failure"):
        await service._analyze_workbook(
            task["task_id"],
            AnalyzeTaskRequest(crawl_task_id=task["task_id"]),
            service._workbook(task["task_id"]),
        )


def test_pipeline_request_rejects_custom_keywords_for_all_sources() -> None:
    with pytest.raises(ValueError, match="data_source=all 时不能传 keywords"):
        PipelineTaskRequest(data_source="all", keywords=["大健康"])


def test_next_daily_run_uses_shanghai_time(tmp_path: Path, monkeypatch) -> None:
    monkeypatch.setattr(
        "douhot_crawler.api.service.chromium_status", lambda: (True, "test browser")
    )
    configured = settings(tmp_path)
    configured.daily_time = "03:00"
    service = ApiTaskService(configured, client=FakeExternalClient())
    timezone = ZoneInfo("Asia/Shanghai")

    same_day = service.next_daily_run(datetime(2026, 8, 4, 2, 30, tzinfo=timezone))
    next_day = service.next_daily_run(datetime(2026, 8, 4, 3, 0, tzinfo=timezone))

    assert same_day == datetime(2026, 8, 4, 3, 0, tzinfo=timezone)
    assert next_day == datetime(2026, 8, 5, 3, 0, tzinfo=timezone)


def test_builtin_scheduler_defaults_to_all_and_skips_any_active_pipeline(
    tmp_path: Path, monkeypatch
) -> None:
    monkeypatch.setattr(
        "douhot_crawler.api.service.chromium_status", lambda: (True, "test browser")
    )
    service = ApiTaskService(settings(tmp_path), client=FakeExternalClient())

    scheduled, created = service._create_daily_pipeline()
    repeated, repeated_created = service._create_daily_pipeline()

    assert created is True
    assert scheduled["params"]["data_source"] == "all"
    assert repeated_created is False
    assert repeated["task_id"] == scheduled["task_id"]


@pytest.mark.asyncio
async def test_service_starts_builtin_daily_scheduler(
    tmp_path: Path, monkeypatch
) -> None:
    monkeypatch.setattr(
        "douhot_crawler.api.service.chromium_status", lambda: (True, "test browser")
    )
    configured = settings(tmp_path)
    configured.daily_enabled = True
    configured.daily_time = "03:00"
    service = ApiTaskService(configured, client=FakeExternalClient())

    await service.start()
    await asyncio.sleep(0)
    health = service.health()
    await service.close()

    assert health["scheduler_enabled"] is True
    assert health["scheduler_time"] == "03:00"
    assert health["scheduler_timezone"] == "Asia/Shanghai"
    assert health["scheduler_next_run_at"] is not None


def test_task_store_fifo_pause_resume_and_restart_recovery(tmp_path: Path) -> None:
    path = tmp_path / "tasks.sqlite3"
    store = ApiTaskStore(path)
    first = store.create(TaskKind.CRAWL, {"keyword": "甲"})
    second = store.create(TaskKind.PIPELINE, {})

    assert store.claim_next()["task_id"] == first["task_id"]
    store.request_pause(first["task_id"])
    assert store.get(first["task_id"])["status"] == TaskStatus.PAUSING
    store.mark_paused(first["task_id"], "user")
    store.resume(first["task_id"])
    assert store.get(first["task_id"])["status"] == TaskStatus.QUEUED

    # FIFO preserves the older resumed task ahead of the later pipeline.
    assert store.claim_next()["task_id"] == first["task_id"]
    restarted = ApiTaskStore(path)
    assert restarted.get(first["task_id"])["status"] == TaskStatus.FAILED
    assert restarted.claim_next()["task_id"] == second["task_id"]
    restarted_again = ApiTaskStore(path)
    assert restarted_again.get(second["task_id"])["status"] == TaskStatus.QUEUED
    assert restarted_again.get(second["task_id"])["pause_reason"] == "shutdown"


def test_pipeline_checkpoints_keep_source_and_sheet_separate(tmp_path: Path) -> None:
    store = ApiTaskStore(tmp_path / "tasks.sqlite3")
    task = store.create(TaskKind.PIPELINE, {"data_source": "all"})

    store.set_pipeline_keywords(
        task["task_id"],
        [
            {"source": "hotspot", "keyword": "大健康", "sheet_name": "大健康"},
            {
                "source": "industry",
                "keyword": "大健康",
                "sheet_name": "行业_大健康_123456",
            },
        ],
    )

    checkpoints = store.pipeline_keywords(task["task_id"])
    assert [(row["source"], row["keyword"], row["sheet_name"]) for row in checkpoints] == [
        ("hotspot", "大健康", "大健康"),
        ("industry", "大健康", "行业_大健康_123456"),
    ]


def test_old_pipeline_checkpoint_schema_migrates_to_hotspot(tmp_path: Path) -> None:
    path = tmp_path / "tasks.sqlite3"
    store = ApiTaskStore(path)
    task = store.create(TaskKind.PIPELINE, {"keywords": ["旧关键词"]})
    with sqlite3.connect(path) as connection:
        connection.execute("DROP TABLE pipeline_keywords")
        connection.execute(
            """
            CREATE TABLE pipeline_keywords (
                task_id TEXT NOT NULL REFERENCES tasks(id) ON DELETE CASCADE,
                position INTEGER NOT NULL,
                keyword TEXT NOT NULL,
                crawl_done INTEGER NOT NULL DEFAULT 0,
                analyze_done INTEGER NOT NULL DEFAULT 0,
                upload_done INTEGER NOT NULL DEFAULT 0,
                last_error TEXT,
                PRIMARY KEY(task_id, position)
            )
            """
        )
        connection.execute(
            "INSERT INTO pipeline_keywords(task_id, position, keyword) "
            "VALUES (?, 0, ?)",
            (task["task_id"], "旧关键词"),
        )

    migrated = ApiTaskStore(path).pipeline_keywords(task["task_id"])[0]
    assert migrated["source"] == "hotspot"
    assert migrated["sheet_name"] == "旧关键词"


def test_active_pipeline_only_reuses_identical_request(
    tmp_path: Path, monkeypatch
) -> None:
    monkeypatch.setattr(
        "douhot_crawler.api.service.chromium_status", lambda: (True, "test browser")
    )
    service = ApiTaskService(settings(tmp_path), client=FakeExternalClient())
    request = PipelineTaskRequest(data_source="hotspot", keywords=["大健康"])
    first, created = service.create_pipeline(request)
    repeated, repeated_created = service.create_pipeline(request)

    assert created is True
    assert repeated_created is False
    assert repeated["task_id"] == first["task_id"]
    with pytest.raises(ApiError, match="已有参数不同的活动流水线"):
        service.create_pipeline(PipelineTaskRequest(data_source="industry"))


def test_active_legacy_pipeline_without_source_is_treated_as_hotspot(
    tmp_path: Path, monkeypatch
) -> None:
    monkeypatch.setattr(
        "douhot_crawler.api.service.chromium_status", lambda: (True, "test browser")
    )
    service = ApiTaskService(settings(tmp_path), client=FakeExternalClient())
    legacy_request = PipelineTaskRequest(
        data_source="hotspot", keywords=["旧关键词"]
    ).model_dump(mode="json")
    legacy_request.pop("data_source")
    existing = service.store.create(TaskKind.PIPELINE, legacy_request)

    repeated, created = service.create_pipeline(
        PipelineTaskRequest(data_source="hotspot", keywords=["旧关键词"])
    )

    assert created is False
    assert repeated["task_id"] == existing["task_id"]


@pytest.mark.asyncio
async def test_external_client_extracts_unique_keywords_and_cookie(tmp_path: Path) -> None:
    calls: list[tuple[str, Any]] = []

    async def handler(request: httpx.Request) -> httpx.Response:
        payload = __import__("json").loads(request.content)
        calls.append((request.url.path, payload))
        if request.url.path == "/hotspots":
            return httpx.Response(
                200,
                json={
                    "code": 200,
                    "message": "操作成功",
                    "data": {
                        "records": [
                            {"title": " 甲 "},
                            {"title": "甲"},
                            {"title": None},
                            {"title": "乙"},
                        ]
                    },
                },
            )
        return httpx.Response(
            200, json={"code": 200, "message": "操作成功", "cookie": "sid=secret"}
        )

    http_client = httpx.AsyncClient(transport=httpx.MockTransport(handler))
    client = ExternalApiClient(settings(tmp_path), client=http_client)
    assert await client.fetch_keywords() == ["甲", "乙"]
    assert await client.fetch_cookie(0) == "sid=secret"
    assert calls == [
        ("/hotspots", {"openId": "test-open-id", "size": 30}),
        ("/cookies", {"type": 0}),
    ]
    await http_client.aclose()


@pytest.mark.asyncio
async def test_external_client_fetches_and_uploads_industry_records(
    tmp_path: Path,
) -> None:
    calls: list[tuple[str, Any]] = []

    async def handler(request: httpx.Request) -> httpx.Response:
        payload = __import__("json").loads(request.content)
        calls.append((request.url.path, payload))
        if request.url.path == "/industry":
            return httpx.Response(
                200,
                json={
                    "code": 200,
                    "message": "操作成功",
                    "data": {
                        "records": [
                            {"title": " 大健康 "},
                            {"title": "大健康"},
                            {"title": "美容"},
                        ]
                    },
                },
            )
        return httpx.Response(200, json={"code": 200, "message": "操作成功"})

    http_client = httpx.AsyncClient(transport=httpx.MockTransport(handler))
    client = ExternalApiClient(settings(tmp_path), client=http_client)
    assert await client.fetch_industry_keywords() == ["大健康", "美容"]
    payload = [{"keyword": "大健康", "videoUrl": "https://example.test/video/1"}]
    await client.upload_industry_rankings(payload)
    assert calls == [
        ("/industry", {"openId": "test-open-id", "size": 30}),
        ("/industry-rankings", payload),
    ]
    await http_client.aclose()


@pytest.mark.asyncio
async def test_external_client_accepts_cookie_in_data_field(tmp_path: Path) -> None:
    async def handler(request: httpx.Request) -> httpx.Response:
        return httpx.Response(
            200,
            json={"code": 200, "message": "操作成功", "data": "sid=from-data"},
        )

    http_client = httpx.AsyncClient(transport=httpx.MockTransport(handler))
    client = ExternalApiClient(settings(tmp_path), client=http_client)
    assert await client.fetch_cookie(0) == "sid=from-data"
    await http_client.aclose()


@pytest.mark.asyncio
async def test_external_client_preserves_hotspot_business_error(tmp_path: Path) -> None:
    async def handler(request: httpx.Request) -> httpx.Response:
        return httpx.Response(200, json={"code": 400, "message": "openId无效"})

    http_client = httpx.AsyncClient(transport=httpx.MockTransport(handler))
    client = ExternalApiClient(settings(tmp_path), client=http_client)
    with pytest.raises(ExternalServiceError, match="openId无效"):
        await client.fetch_keywords()
    await http_client.aclose()


@pytest.mark.asyncio
async def test_external_client_retries_5xx_three_times(tmp_path: Path, monkeypatch) -> None:
    attempts = 0

    async def handler(request: httpx.Request) -> httpx.Response:
        nonlocal attempts
        attempts += 1
        if attempts < 4:
            return httpx.Response(503)
        return httpx.Response(200, json={"code": 200, "message": "ok"})

    async def no_sleep(delay: float) -> None:
        return None

    monkeypatch.setattr("douhot_crawler.api.clients.asyncio.sleep", no_sleep)
    http_client = httpx.AsyncClient(transport=httpx.MockTransport(handler))
    client = ExternalApiClient(settings(tmp_path), client=http_client)
    await client.upload_rankings([{"videoUrl": "https://example.test/video/1"}])
    assert attempts == 4
    await http_client.aclose()


@pytest.mark.asyncio
async def test_external_client_retries_business_5xx_three_times(
    tmp_path: Path, monkeypatch
) -> None:
    attempts = 0

    async def handler(request: httpx.Request) -> httpx.Response:
        nonlocal attempts
        attempts += 1
        if attempts < 4:
            return httpx.Response(
                200, json={"code": 500, "message": "请求处理异常，请稍后再试"}
            )
        return httpx.Response(
            200, json={"code": 200, "message": "操作成功", "data": "sid=ok"}
        )

    async def no_sleep(delay: float) -> None:
        return None

    monkeypatch.setattr("douhot_crawler.api.clients.asyncio.sleep", no_sleep)
    http_client = httpx.AsyncClient(transport=httpx.MockTransport(handler))
    client = ExternalApiClient(settings(tmp_path), client=http_client)
    assert await client.fetch_cookie(0) == "sid=ok"
    assert attempts == 4
    await http_client.aclose()


@pytest.mark.asyncio
async def test_external_client_posts_callback_array_and_accepts_empty_204(
    tmp_path: Path,
) -> None:
    observed: dict[str, Any] = {}

    async def handler(request: httpx.Request) -> httpx.Response:
        observed["body"] = __import__("json").loads(request.content)
        observed["task_id"] = request.headers["X-DouHot-Task-ID"]
        return httpx.Response(204)

    http_client = httpx.AsyncClient(transport=httpx.MockTransport(handler))
    client = ExternalApiClient(settings(tmp_path), client=http_client)
    await client.send_callback(
        "https://callback.example.test/result",
        [{"type": 0, "keyword": "大健康"}],
        task_id="42118d44-6334-4a0c-a9a5-9a5096ab2962",
    )

    assert observed == {
        "body": [{"type": 0, "keyword": "大健康"}],
        "task_id": "42118d44-6334-4a0c-a9a5-9a5096ab2962",
    }
    await http_client.aclose()


@pytest.mark.asyncio
async def test_external_client_rejects_redirecting_callback(tmp_path: Path) -> None:
    async def handler(request: httpx.Request) -> httpx.Response:
        return httpx.Response(302, headers={"location": "http://127.0.0.1/private"})

    http_client = httpx.AsyncClient(transport=httpx.MockTransport(handler))
    client = ExternalApiClient(settings(tmp_path), client=http_client)

    with pytest.raises(ExternalServiceError, match="HTTP 302"):
        await client.send_callback(
            "https://callback.example.test/result",
            [],
            task_id="42118d44-6334-4a0c-a9a5-9a5096ab2962",
        )
    await http_client.aclose()


def test_follower_count_parser() -> None:
    assert parse_follower_count("1.2万") == (12_000, False)
    assert parse_follower_count("3.45亿") == (345_000_000, False)
    assert parse_follower_count("12,345") == (12_345, False)
    assert parse_follower_count(12.6) == (13, False)
    assert parse_follower_count("") == (0, True)
    assert parse_follower_count("未知") == (0, True)


class FakeExternalClient:
    def __init__(self) -> None:
        self.uploads: list[list[dict[str, Any]]] = []
        self.industry_uploads: list[list[dict[str, Any]]] = []
        self.callbacks: list[tuple[str, list[dict[str, Any]]]] = []

    async def fetch_keywords(self) -> list[str]:
        return ["甲", "乙"]

    async def fetch_industry_keywords(self) -> list[str]:
        return ["甲", "乙"]

    async def fetch_cookie(self, cookie_type: int) -> str:
        return f"type={cookie_type}"

    async def upload_rankings(self, records: list[dict[str, Any]]) -> None:
        self.uploads.append(records)

    async def upload_industry_rankings(
        self, records: list[dict[str, Any]]
    ) -> None:
        self.industry_uploads.append(records)

    async def send_callback(
        self,
        callback_url: str,
        records: list[dict[str, Any]],
        *,
        task_id: str,
    ) -> None:
        self.callbacks.append((callback_url, records))

    async def close(self) -> None:
        return None


def create_upload_workbook(
    path: Path, *, rows: int, include_play_url: bool = False
) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "关键词"
    headers = [*RESULT_HEADERS, "视频口播"]
    if include_play_url:
        headers.append("视频播放地址")
    sheet.append(headers)
    for index in range(rows):
        row = [
            index + 1,
            "低粉爆款",
            "2026-08-04 03:00:00",
            "近7天",
            f"视频{index}",
            f"https://www.douyin.com/video/{index}",
            f"作者{index}",
            "1万",
            "50万",
            "20万",
            "1万",
            "5%",
            "好评",
            "口播",
        ]
        if include_play_url:
            row.append(f"https://aweme.snssdk.com/aweme/v1/play/?video_id={index}")
        sheet.append(row)
    atomic_workbook_save(workbook, path)
    workbook.close()


def test_collect_endpoint_waits_and_returns_ranking_payload_without_callback(
    tmp_path: Path, monkeypatch
) -> None:
    monkeypatch.setattr(
        "douhot_crawler.api.service.chromium_status", lambda: (True, "test browser")
    )

    async def fake_crawler(options, **kwargs):
        path = Path(kwargs["excel_path"])
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = excel_sheet_name(options.keyword)
        sheet.append(RESULT_HEADERS)
        sheet.append(
            [
                1,
                options.result_type,
                "2026-08-07 03:00:00",
                options.time_range,
                "健康视频",
                "https://www.douyin.com/video/1",
                "健康作者",
                "1.2万",
                "50万",
                "20万",
                "1万",
                "5%",
                "值得收藏",
            ]
        )
        atomic_workbook_save(workbook, path)
        workbook.close()
        return {
            "excel_path": str(path),
            "added_count": 1,
            "skipped_count": 0,
            "sheet": excel_sheet_name(options.keyword),
            "stopped": False,
        }

    def fake_analyze(args, *, stop_requested):
        workbook = load_workbook(args.excel)
        sheet = workbook[args.sheet[0]]
        sheet.cell(row=1, column=14, value="视频口播")
        sheet.cell(row=1, column=15, value="视频播放地址")
        sheet.cell(row=2, column=14, value="健康测试口播")
        sheet.cell(
            row=2,
            column=15,
            value="https://aweme.snssdk.com/aweme/v1/play/?video_id=1",
        )
        atomic_workbook_save(workbook, args.excel)
        workbook.close()
        return 1, 0, 0

    monkeypatch.setattr("douhot_crawler.api.service.run_crawler", fake_crawler)
    monkeypatch.setattr("douhot_crawler.api.service.analyze_excel", fake_analyze)
    fake_client = FakeExternalClient()
    service = ApiTaskService(settings(tmp_path), client=fake_client)

    with TestClient(create_app(settings(tmp_path), service=service)) as client:
        response = client.post(
            "/api/v1/viral-videos/collect",
            json={"keyword": "大健康/创业"},
        )

    assert response.status_code == 200
    assert response.json() == [
        {
            "type": 0,
            "keyword": "大健康/创业",
            "videoName": "健康视频",
            "videoUrl": "https://www.douyin.com/video/1",
            "authorName": "健康作者",
            "followerCount": 12000,
            "heatValue": "50万",
            "newPlayCount": "20万",
            "newLikeCount": "1万",
            "likeRate": "5%",
            "highPraiseComment": "值得收藏",
            "videoOral": "健康测试口播",
            "videoPlayUrl": (
                "https://aweme.snssdk.com/aweme/v1/play/?video_id=1"
            ),
        }
    ]
    assert fake_client.uploads == []


def test_collect_endpoint_returns_empty_array_when_keyword_has_no_videos(
    tmp_path: Path, monkeypatch
) -> None:
    monkeypatch.setattr(
        "douhot_crawler.api.service.chromium_status", lambda: (True, "test browser")
    )

    async def fake_crawler(options, **kwargs):
        return {
            "excel_path": str(kwargs["excel_path"]),
            "added_count": 0,
            "skipped_count": 0,
            "sheet": excel_sheet_name(options.keyword),
            "stopped": False,
        }

    monkeypatch.setattr("douhot_crawler.api.service.run_crawler", fake_crawler)
    service = ApiTaskService(settings(tmp_path), client=FakeExternalClient())

    with TestClient(create_app(settings(tmp_path), service=service)) as client:
        response = client.post(
            "/api/v1/viral-videos/collect",
            json={"keyword": "没有结果的关键词"},
        )

    assert response.status_code == 200
    assert response.json() == []


def test_collect_endpoint_returns_accepted_and_posts_result_to_callback(
    tmp_path: Path, monkeypatch
) -> None:
    monkeypatch.setattr(
        "douhot_crawler.api.service.chromium_status", lambda: (True, "test browser")
    )

    async def fake_crawler(options, **kwargs):
        path = Path(kwargs["excel_path"])
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = options.keyword
        sheet.append(RESULT_HEADERS)
        sheet.append(
            [
                1,
                options.result_type,
                "2026-08-07 03:00:00",
                options.time_range,
                "回调视频",
                "https://www.douyin.com/video/2",
                "回调作者",
                "2万",
                "60万",
                "30万",
                "2万",
                "6%",
                "回调评论",
            ]
        )
        atomic_workbook_save(workbook, path)
        workbook.close()
        return {
            "excel_path": str(path),
            "added_count": 1,
            "skipped_count": 0,
            "sheet": options.keyword,
            "stopped": False,
        }

    def fake_analyze(args, *, stop_requested):
        workbook = load_workbook(args.excel)
        sheet = workbook[args.sheet[0]]
        sheet.cell(row=1, column=14, value="视频口播")
        sheet.cell(row=1, column=15, value="视频播放地址")
        sheet.cell(row=2, column=14, value="异步回调口播")
        sheet.cell(
            row=2,
            column=15,
            value="https://aweme.snssdk.com/aweme/v1/play/?video_id=2",
        )
        atomic_workbook_save(workbook, args.excel)
        workbook.close()
        return 1, 0, 0

    monkeypatch.setattr("douhot_crawler.api.service.run_crawler", fake_crawler)
    monkeypatch.setattr("douhot_crawler.api.service.analyze_excel", fake_analyze)
    fake_client = FakeExternalClient()
    service = ApiTaskService(settings(tmp_path), client=fake_client)
    expected = [
        {
            "type": 0,
            "keyword": "大健康",
            "videoName": "回调视频",
            "videoUrl": "https://www.douyin.com/video/2",
            "authorName": "回调作者",
            "followerCount": 20000,
            "heatValue": "60万",
            "newPlayCount": "30万",
            "newLikeCount": "2万",
            "likeRate": "6%",
            "highPraiseComment": "回调评论",
            "videoOral": "异步回调口播",
            "videoPlayUrl": (
                "https://aweme.snssdk.com/aweme/v1/play/?video_id=2"
            ),
        }
    ]

    with TestClient(create_app(settings(tmp_path), service=service)) as client:
        response = client.post(
            "/api/v1/viral-videos/collect",
            json={
                "keyword": "大健康",
                "callback_url": "https://callback.example.test/result",
            },
        )
        assert response.status_code == 202
        accepted = response.json()
        assert accepted["status"] == "queued"
        assert accepted["created"] is True

        for _ in range(100):
            task = client.get(f"/api/v1/tasks/{accepted['task_id']}").json()
            if task["status"] in {"succeeded", "succeeded_with_warnings", "failed"}:
                break
            time.sleep(0.01)

    assert task["status"] == "succeeded"
    assert fake_client.callbacks == [
        ("https://callback.example.test/result", expected)
    ]
    assert fake_client.uploads == []


def test_collect_endpoint_reports_failure_when_all_crawled_videos_fail_analysis(
    tmp_path: Path, monkeypatch
) -> None:
    monkeypatch.setattr(
        "douhot_crawler.api.service.chromium_status", lambda: (True, "test browser")
    )

    async def fake_crawler(options, **kwargs):
        path = Path(kwargs["excel_path"])
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = options.keyword
        sheet.append(RESULT_HEADERS)
        sheet.append(
            [
                1,
                options.result_type,
                "2026-08-07 03:00:00",
                options.time_range,
                "失败视频",
                "https://www.douyin.com/video/failed",
                "失败作者",
                "1万",
                "50万",
                "20万",
                "1万",
                "5%",
                "",
            ]
        )
        atomic_workbook_save(workbook, path)
        workbook.close()
        return {
            "excel_path": str(path),
            "added_count": 1,
            "skipped_count": 0,
            "sheet": options.keyword,
            "stopped": False,
        }

    def fake_analyze(args, *, stop_requested):
        workbook = load_workbook(args.excel)
        sheet = workbook[args.sheet[0]]
        sheet.cell(row=1, column=14, value="视频口播")
        sheet.cell(row=1, column=15, value="视频播放地址")
        atomic_workbook_save(workbook, args.excel)
        workbook.close()
        return 0, 0, 1

    monkeypatch.setattr("douhot_crawler.api.service.run_crawler", fake_crawler)
    monkeypatch.setattr("douhot_crawler.api.service.analyze_excel", fake_analyze)
    service = ApiTaskService(settings(tmp_path), client=FakeExternalClient())

    with TestClient(create_app(settings(tmp_path), service=service)) as client:
        response = client.post(
            "/api/v1/viral-videos/collect",
            json={"keyword": "大健康"},
        )

    assert response.status_code == 502
    assert response.json()["error"] == {
        "code": "COLLECT_FAILED",
        "message": "所有已爬取视频的口播解析均失败",
        "details": None,
    }


@pytest.mark.asyncio
async def test_failed_upload_batch_is_resumable(tmp_path: Path, monkeypatch) -> None:
    monkeypatch.setattr(
        "douhot_crawler.api.service.chromium_status", lambda: (True, "test browser")
    )

    class FailOnceClient(FakeExternalClient):
        def __init__(self) -> None:
            super().__init__()
            self.fail = True

        async def upload_rankings(self, records: list[dict[str, Any]]) -> None:
            self.uploads.append(records)
            if self.fail:
                self.fail = False
                raise RuntimeError("temporary upload failure")

    fake_client = FailOnceClient()
    service = ApiTaskService(settings(tmp_path), client=fake_client)
    task, _ = service.create_pipeline(
        PipelineTaskRequest(data_source="hotspot", keywords=["关键词"])
    )
    workbook = service._workbook(task["task_id"])
    workbook.parent.mkdir(parents=True, exist_ok=True)
    create_upload_workbook(workbook, rows=2)

    with pytest.raises(TaskPaused, match="upload_failure"):
        await service._upload_sheet(task["task_id"], workbook, "关键词")
    await service._upload_sheet(task["task_id"], workbook, "关键词")
    assert [len(batch) for batch in fake_client.uploads] == [2, 2]
    assert all(
        row["videoPlayUrl"] == ""
        for batch in fake_client.uploads
        for row in batch
    )


@pytest.mark.asyncio
async def test_industry_upload_uses_separate_target_and_idempotency_key(
    tmp_path: Path, monkeypatch
) -> None:
    monkeypatch.setattr(
        "douhot_crawler.api.service.chromium_status", lambda: (True, "test browser")
    )
    fake_client = FakeExternalClient()
    service = ApiTaskService(settings(tmp_path), client=fake_client)
    task = service.store.create(TaskKind.PIPELINE, {"data_source": "industry"})
    workbook = service._workbook(task["task_id"])
    workbook.parent.mkdir(parents=True, exist_ok=True)
    create_upload_workbook(workbook, rows=1)

    await service._upload_sheet(
        task["task_id"],
        workbook,
        "关键词",
        keyword="大健康",
        source="industry",
    )
    await service._upload_sheet(
        task["task_id"],
        workbook,
        "关键词",
        keyword="大健康",
        source="industry",
    )

    assert fake_client.uploads == []
    assert len(fake_client.industry_uploads) == 1
    assert fake_client.industry_uploads[0][0]["keyword"] == "大健康"


@pytest.mark.asyncio
async def test_standalone_upload_task_sends_existing_workbook(
    tmp_path: Path, monkeypatch
) -> None:
    monkeypatch.setattr(
        "douhot_crawler.api.service.chromium_status", lambda: (True, "test browser")
    )
    fake_client = FakeExternalClient()
    service = ApiTaskService(settings(tmp_path), client=fake_client)
    source = service.store.create(TaskKind.CRAWL, {"keyword": "关键词"})
    workbook = service._workbook(source["task_id"])
    workbook.parent.mkdir(parents=True, exist_ok=True)
    create_upload_workbook(workbook, rows=21, include_play_url=True)
    service.store.finish(
        source["task_id"],
        result={"artifact": service._artifact(workbook)},
        artifact_path=workbook,
    )

    upload = service.create_upload(
        UploadTaskRequest(source_task_id=source["task_id"], sheets=None)
    )
    await service._execute(service.store.claim_next())

    completed = service.status(upload["task_id"])
    assert completed["status"] == TaskStatus.SUCCEEDED
    assert completed["result"]["sheets"] == ["关键词"]
    assert completed["result"]["eligible_rows"] == 21
    assert completed["result"]["sent_rows"] == 21
    assert [len(batch) for batch in fake_client.uploads] == [20, 1]
    assert fake_client.uploads[0][0]["videoPlayUrl"] == (
        "https://aweme.snssdk.com/aweme/v1/play/?video_id=0"
    )


@pytest.mark.asyncio
async def test_pipeline_runs_keywords_sequentially_and_uploads_batches(
    tmp_path: Path, monkeypatch
) -> None:
    monkeypatch.setattr(
        "douhot_crawler.api.service.chromium_status", lambda: (True, "test browser")
    )
    fake_client = FakeExternalClient()
    service = ApiTaskService(settings(tmp_path), client=fake_client)
    crawl_order: list[tuple[str, str]] = []
    analyze_order: list[str] = []
    crawl_limits: list[int] = []
    analyze_success_limits: list[int] = []

    async def fake_crawler(options, **kwargs):
        storage_keyword = kwargs.get("storage_keyword", options.keyword)
        crawl_order.append((options.keyword, storage_keyword))
        crawl_limits.append(kwargs["max_results"])
        path = Path(kwargs["excel_path"])
        if path.exists():
            workbook = load_workbook(path)
            sheet = workbook.create_sheet(storage_keyword)
        else:
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = storage_keyword
        sheet.append(RESULT_HEADERS)
        for index in range(kwargs["max_results"]):
            sheet.append(
                [
                    index + 1,
                    options.result_type,
                    "2026-08-04 03:00:00",
                    options.time_range,
                    f"视频{index}",
                    f"https://www.douyin.com/video/{options.keyword}{index}",
                    f"作者{index}",
                    "1.2万" if index else "未知",
                    "50万",
                    "20万",
                    "1万",
                    "5%",
                    "好评",
                ]
            )
        atomic_workbook_save(workbook, path)
        workbook.close()
        return {
            "excel_path": str(path),
            "added_count": kwargs["max_results"],
            "skipped_count": 0,
            "sheet": storage_keyword,
            "stopped": False,
        }

    def fake_analyze(args, *, stop_requested):
        analyze_order.append(args.sheet[0])
        analyze_success_limits.append(args.success_limit)
        workbook = load_workbook(args.excel)
        sheet = workbook[args.sheet[0]]
        sheet.cell(row=1, column=14, value="视频口播")
        sheet.cell(row=1, column=15, value="视频播放地址")
        # Simulate two failed candidates followed by the three required transcripts.
        for row in range(4, 4 + args.success_limit):
            sheet.cell(row=row, column=14, value="测试口播")
            sheet.cell(
                row=row,
                column=15,
                value=f"https://aweme.snssdk.com/aweme/v1/play/?row={row}",
            )
        atomic_workbook_save(workbook, args.excel)
        workbook.close()
        return args.success_limit, 0, 2

    monkeypatch.setattr("douhot_crawler.api.service.run_crawler", fake_crawler)
    monkeypatch.setattr("douhot_crawler.api.service.analyze_excel", fake_analyze)

    task, created = service.create_pipeline(PipelineTaskRequest())
    assert created is True
    claimed = service.store.claim_next()
    await service._execute(claimed)

    completed = service.status(task["task_id"])
    assert completed["status"] == TaskStatus.SUCCEEDED_WITH_WARNINGS
    assert completed["result"]["keywords_succeeded"] == 4
    assert completed["result"]["sources"] == {
        "hotspot": {"total": 2, "succeeded": 2, "failed": 0},
        "industry": {"total": 2, "succeeded": 2, "failed": 0},
    }
    assert [keyword for keyword, _ in crawl_order] == ["甲", "乙", "甲", "乙"]
    assert [sheet for _, sheet in crawl_order[:2]] == ["甲", "乙"]
    assert all(sheet.startswith("行业_") for _, sheet in crawl_order[2:])
    assert len(set(analyze_order)) == 4
    assert crawl_limits == [15, 15, 15, 15]
    assert analyze_success_limits == [3, 3, 3, 3]
    assert [len(batch) for batch in fake_client.uploads] == [3, 3]
    assert [len(batch) for batch in fake_client.industry_uploads] == [3, 3]
    assert all(
        isinstance(row["followerCount"], int)
        for batch in fake_client.uploads
        for row in batch
    )
    assert all(
        row["videoPlayUrl"].startswith("https://aweme.snssdk.com/aweme/v1/play/")
        for batch in fake_client.uploads
        for row in batch
    )


@pytest.mark.asyncio
@pytest.mark.parametrize(
    ("valid_transcripts", "expected_upload_sizes"),
    [(2, [2]), (0, [])],
)
async def test_pipeline_uploads_partial_result_when_transcript_target_not_reached(
    tmp_path: Path,
    monkeypatch,
    valid_transcripts: int,
    expected_upload_sizes: list[int],
) -> None:
    monkeypatch.setattr(
        "douhot_crawler.api.service.chromium_status", lambda: (True, "test browser")
    )

    async def fake_crawler(options, **kwargs):
        path = Path(kwargs["excel_path"])
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = kwargs["storage_keyword"]
        sheet.append(RESULT_HEADERS)
        for index in range(kwargs["max_results"]):
            sheet.append(
                [
                    index + 1,
                    options.result_type,
                    "2026-08-07 03:00:00",
                    options.time_range,
                    f"视频{index}",
                    f"https://www.douyin.com/video/{index}",
                    f"作者{index}",
                    "1万",
                    "50万",
                    "20万",
                    "1万",
                    "5%",
                    "好评",
                ]
            )
        atomic_workbook_save(workbook, path)
        workbook.close()
        return {
            "excel_path": str(path),
            "added_count": kwargs["max_results"],
            "skipped_count": 0,
            "sheet": kwargs["storage_keyword"],
            "stopped": False,
        }

    def fake_analyze(args, *, stop_requested):
        assert args.success_limit == 3
        workbook = load_workbook(args.excel)
        sheet = workbook[args.sheet[0]]
        sheet.cell(row=1, column=14, value="视频口播")
        for index in range(valid_transcripts):
            sheet.cell(row=index + 2, column=14, value=f"口播{index + 1}")
        atomic_workbook_save(workbook, args.excel)
        workbook.close()
        return valid_transcripts, 0, 5 - valid_transcripts

    monkeypatch.setattr("douhot_crawler.api.service.run_crawler", fake_crawler)
    monkeypatch.setattr("douhot_crawler.api.service.analyze_excel", fake_analyze)
    fake_client = FakeExternalClient()
    service = ApiTaskService(settings(tmp_path), client=fake_client)
    task, _ = service.create_pipeline(
        PipelineTaskRequest(
            data_source="hotspot",
            keywords=["大健康"],
            limit_per_keyword=3,
            candidate_limit_per_keyword=5,
        )
    )

    await service._execute(service.store.claim_next())

    completed = service.status(task["task_id"])
    assert completed["status"] == TaskStatus.SUCCEEDED_WITH_WARNINGS
    assert completed["progress"]["valid_count"] == valid_transcripts
    assert any(
        "TARGET_NOT_REACHED" in warning
        for warning in completed["progress"]["warnings"]
    )
    assert [len(batch) for batch in fake_client.uploads] == expected_upload_sizes


@pytest.mark.asyncio
async def test_pipeline_resume_counts_existing_transcripts_toward_target(
    tmp_path: Path, monkeypatch
) -> None:
    monkeypatch.setattr(
        "douhot_crawler.api.service.chromium_status", lambda: (True, "test browser")
    )

    async def unexpected_crawl(*args, **kwargs):
        raise AssertionError("恢复后的 crawl_done 检查点不应重新爬取")

    observed_success_limits: list[int] = []

    def fake_analyze(args, *, stop_requested):
        observed_success_limits.append(args.success_limit)
        workbook = load_workbook(args.excel)
        sheet = workbook[args.sheet[0]]
        sheet.cell(row=4, column=14, value="恢复后新增口播")
        atomic_workbook_save(workbook, args.excel)
        workbook.close()
        return 1, 2, 0

    monkeypatch.setattr("douhot_crawler.api.service.run_crawler", unexpected_crawl)
    monkeypatch.setattr("douhot_crawler.api.service.analyze_excel", fake_analyze)
    fake_client = FakeExternalClient()
    service = ApiTaskService(settings(tmp_path), client=fake_client)
    request = PipelineTaskRequest(
        data_source="hotspot",
        keywords=["大健康"],
        limit_per_keyword=3,
        candidate_limit_per_keyword=5,
    )
    task, _ = service.create_pipeline(request)
    service.store.set_pipeline_keywords(
        task["task_id"],
        [{"source": "hotspot", "keyword": "大健康", "sheet_name": "大健康"}],
    )
    service.store.update_pipeline_keyword(task["task_id"], 0, crawl_done=True)

    workbook_path = service._workbook(task["task_id"])
    workbook_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "大健康"
    sheet.append([*RESULT_HEADERS, "视频口播"])
    for index in range(5):
        sheet.append(
            [
                index + 1,
                "低粉爆款",
                "2026-08-07 03:00:00",
                "近7天",
                f"视频{index}",
                f"https://www.douyin.com/video/{index}",
                f"作者{index}",
                "1万",
                "50万",
                "20万",
                "1万",
                "5%",
                "好评",
                "已有口播" if index < 2 else "",
            ]
        )
    atomic_workbook_save(workbook, workbook_path)
    workbook.close()

    await service._execute(service.store.claim_next())

    completed = service.status(task["task_id"])
    assert completed["status"] == TaskStatus.SUCCEEDED
    assert observed_success_limits == [1]
    assert completed["progress"]["valid_count"] == 3
    assert [len(batch) for batch in fake_client.uploads] == [3]


@pytest.mark.asyncio
async def test_pipeline_upload_hard_caps_eligible_rows(tmp_path: Path, monkeypatch) -> None:
    monkeypatch.setattr(
        "douhot_crawler.api.service.chromium_status", lambda: (True, "test browser")
    )
    fake_client = FakeExternalClient()
    service = ApiTaskService(settings(tmp_path), client=fake_client)
    task = service.store.create(TaskKind.PIPELINE, {"data_source": "hotspot"})
    workbook = service._workbook(task["task_id"])
    workbook.parent.mkdir(parents=True, exist_ok=True)
    create_upload_workbook(workbook, rows=8, include_play_url=True)

    counts = await service._upload_sheet(
        task["task_id"],
        workbook,
        "关键词",
        max_records=3,
    )

    assert counts == {"eligible": 3, "sent": 3}
    assert [len(batch) for batch in fake_client.uploads] == [3]


class FakeRouteService:
    async def start(self) -> None:
        return None

    async def close(self) -> None:
        return None

    def health(self) -> dict[str, Any]:
        return {
            "status": "ok",
            "worker_running": True,
            "database_ok": True,
            "browser_ok": True,
            "external_urls_configured": True,
            "scheduler_overlap": False,
            "scheduler_enabled": True,
            "scheduler_time": "03:00",
            "scheduler_timezone": "Asia/Shanghai",
            "scheduler_next_run_at": "2026-08-05T03:00:00+08:00",
        }

    async def keywords(self) -> list[str]:
        return ["关键词"]

    def create_pipeline(self, request):
        return {"task_id": "id", "status": "queued"}, True

    def create_upload(self, request):
        return {"task_id": "upload-id", "status": "queued"}


def test_api_routes_and_validation_envelope(tmp_path: Path) -> None:
    app = create_app(settings(tmp_path), service=FakeRouteService())
    with TestClient(app) as client:
        assert client.get("/api/v1/health").json()["status"] == "ok"
        assert client.get("/api/v1/keywords").json() == {"key_word": ["关键词"]}
        accepted = client.post("/api/v1/tasks/pipeline", json={})
        assert accepted.status_code == 202
        assert accepted.json() == {"task_id": "id", "status": "queued", "created": True}
        invalid = client.post(
            "/api/v1/tasks/pipeline",
            json={"keywords": [f"词{index}" for index in range(31)]},
        )
        assert invalid.status_code == 422
        assert invalid.json()["error"]["code"] == "VALIDATION_ERROR"
        uploaded = client.post(
            "/api/v1/tasks/upload",
            json={
                "source_task_id": "42118d44-6334-4a0c-a9a5-9a5096ab2962",
                "sheets": None,
            },
        )
        assert uploaded.status_code == 202
        assert uploaded.json()["task_id"] == "upload-id"
        missing_keyword = client.post("/api/v1/viral-videos/collect", json={})
        assert missing_keyword.status_code == 422
        invalid_callback = client.post(
            "/api/v1/viral-videos/collect",
            json={"keyword": "大健康", "callback_url": "ftp://example.test/result"},
        )
        assert invalid_callback.status_code == 422


def test_openapi_documents_workflow_examples_and_errors(tmp_path: Path) -> None:
    schema = create_app(settings(tmp_path), service=FakeRouteService()).openapi()

    assert "FIFO" in schema["info"]["description"]
    assert [item["name"] for item in schema["tags"]] == [
        "系统",
        "关键词",
        "任务创建",
        "任务控制",
    ]
    crawl = schema["paths"]["/api/v1/tasks/crawl"]["post"]
    assert crawl["summary"] == "创建单关键词爬取任务"
    assert crawl["tags"] == ["任务创建"]
    assert crawl["responses"]["202"]["description"] == "Successful Response"
    assert "ErrorResponse" in str(crawl["responses"]["502"])
    request_schema = schema["components"]["schemas"]["CrawlTaskRequest"]
    assert request_schema["examples"][0]["keyword"] == "大健康"
    assert "最多采集条数" in request_schema["properties"]["limit"]["description"]
    pipeline_schema = schema["components"]["schemas"]["PipelineTaskRequest"]
    assert pipeline_schema["properties"]["data_source"]["default"] == "all"
    assert "有效口播条数" in pipeline_schema["properties"]["limit_per_keyword"][
        "description"
    ]
    assert "候选视频数" in pipeline_schema["properties"][
        "candidate_limit_per_keyword"
    ]["description"]
    assert pipeline_schema["examples"][0]["data_source"] == "all"
    assert pipeline_schema["examples"][1]["data_source"] == "hotspot"
    pipeline = schema["paths"]["/api/v1/tasks/pipeline"]["post"]
    assert "rankingViralVideoByIndustry" in pipeline["description"]
    assert "TARGET_NOT_REACHED" in pipeline["description"]
    task_schema = schema["components"]["schemas"]["TaskResponse"]
    assert task_schema["examples"][0]["progress"]["current"] == 3
    upload = schema["paths"]["/api/v1/tasks/upload"]["post"]
    assert upload["summary"] == "上传现有 Excel 的全部合格数据"
    collect = schema["paths"]["/api/v1/viral-videos/collect"]["post"]
    assert collect["summary"] == "按关键词爬取并返回榜单视频数据"
    assert collect["responses"]["202"]["description"] == "已创建异步回调任务"
    collect_request = schema["components"]["schemas"]["CollectKeywordRequest"]
    assert collect_request["required"] == ["keyword"]
    assert collect_request["examples"] == [
        {
            "keyword": "大健康",
            "result_type": "低粉爆款",
            "time_range": "近7天",
            "input_timeout": 30,
            "detail_delay": 1,
            "limit": 3,
            "analyze_timeout": 90,
            "analyze_delay": 0,
            "callback_url": "https://example.com/hooks/douhot",
        }
    ]
    assert "异步结果回调地址" in collect_request["properties"]["callback_url"][
        "description"
    ]


def test_daily_launcher_submits_once_and_prints_task_id(monkeypatch, capsys) -> None:
    request: dict[str, Any] = {}

    def fake_post(url: str, *, json: dict[str, Any], timeout: float):
        request.update(url=url, json=json, timeout=timeout)
        return httpx.Response(
            202,
            json={"task_id": "pipeline-id", "status": "queued", "created": True},
            request=httpx.Request("POST", url),
        )

    monkeypatch.setenv("DOUHOT_DAILY_API_URL", "http://127.0.0.1:9999/")
    monkeypatch.setattr(daily.httpx, "post", fake_post)
    daily.main()

    assert request == {
        "url": "http://127.0.0.1:9999/api/v1/tasks/pipeline",
        "json": {"data_source": "all"},
        "timeout": 30.0,
    }
    assert "task_id=pipeline-id status=queued created=true" in capsys.readouterr().out
