import argparse
import asyncio
import re
import sys
from datetime import datetime
from pathlib import Path

from crawl4ai import (AsyncWebCrawler, BrowserConfig, CacheMode,
                      CrawlerRunConfig)
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from playwright.async_api import Locator, Page
from playwright.async_api import TimeoutError as PlaywrightTimeoutError

TARGET_URL = (
    "https://douhot.douyin.com/square/hotspot"
    "?active_tab=hotspot_video"
    "&date_window=168"
    "&sub_type=1001"
)

DEFAULT_RESULT_TYPE = "低粉爆款"
DEFAULT_TIME_RANGE = "近7天"
TIME_RANGE_CHOICES = ("近1小时", "近1天", "近3天", "近7天")
DOUYIN_VIDEO_URL_PREFIX = "https://www.douyin.com/video/"
RESULT_EXCEL_PATH = Path("result") / "result.xlsx"
RESULT_HEADERS = [
    "序号",
    "类型",
    "爬取到的时间",
    "时间类型",
    "视频名称",
    "视频的url",
    "博主名称",
    "总粉丝数",
    "热度值",
    "新增播放量",
    "新增点赞量",
    "点赞率",
]

PROFILE_PATH = Path.home() / ".crawl4ai" / "profiles" / "douhot"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="使用 douhot Profile 搜索热门关键词并爬取结果"
    )

    parser.add_argument(
        "keyword",
        help='热门搜索关键词，例如 "大健康"',
    )

    parser.add_argument(
        "--input-timeout",
        type=float,
        default=30.0,
        help="等待搜索输入框出现的最长秒数（默认：30）",
    )

    parser.add_argument(
        "--result-type",
        default=DEFAULT_RESULT_TYPE,
        help=f"搜索后点击的类型筛选（默认：{DEFAULT_RESULT_TYPE}）",
    )

    parser.add_argument(
        "--time-range",
        choices=TIME_RANGE_CHOICES,
        default=DEFAULT_TIME_RANGE,
        help=f"搜索后点击的时间筛选（默认：{DEFAULT_TIME_RANGE}）",
    )

    parser.add_argument(
        "--headless",
        action="store_true",
        help="无头模式运行；首次调试不要添加",
    )

    return parser.parse_args()


async def print_input_candidates(page: Page) -> None:
    """输出当前页面所有可见输入框，方便定位选择器。"""

    inputs = page.locator("input")
    count = await inputs.count()

    print(f"\n页面共发现 {count} 个 input 元素：")

    for index in range(count):
        element = inputs.nth(index)

        try:
            visible = await element.is_visible()
            editable = await element.is_editable()

            if not visible:
                continue

            input_type = await element.get_attribute("type")
            placeholder = await element.get_attribute("placeholder")
            aria_label = await element.get_attribute("aria-label")
            class_name = await element.get_attribute("class")

            print(
                f"[input {index}] "
                f"editable={editable}, "
                f"type={input_type!r}, "
                f"placeholder={placeholder!r}, "
                f"aria-label={aria_label!r}, "
                f"class={class_name!r}"
            )

        except Exception:
            continue


async def find_search_input(
    page: Page,
) -> Locator:
    """查找最可能的搜索输入框。"""

    selectors = [
        'input[placeholder*="热门搜索"]',
        'input[placeholder*="搜索"]',
        'input[placeholder*="关键词"]',
        'input[aria-label*="搜索"]',
        'input[type="search"]',
    ]

    for selector in selectors:
        locator = page.locator(selector)
        count = await locator.count()

        for index in range(count):
            candidate = locator.nth(index)

            if await candidate.is_visible() and await candidate.is_editable():
                print(f"自动选中搜索框：{selector}")
                return candidate

    # 找不到明确标记时，对可见文本输入框评分。
    candidates: list[tuple[int, Locator, str]] = []

    all_inputs = page.locator(
        'input:not([type]), input[type="text"], input[type="search"]'
    )

    count = await all_inputs.count()

    for index in range(count):
        candidate = all_inputs.nth(index)

        try:
            if not await candidate.is_visible():
                continue

            if not await candidate.is_editable():
                continue

            placeholder = await candidate.get_attribute("placeholder") or ""
            aria_label = await candidate.get_attribute("aria-label") or ""
            input_type = await candidate.get_attribute("type") or "text"

            surrounding_text = await candidate.evaluate(
                """
                element => {
                    const parent =
                        element.closest('form')
                        || element.parentElement?.parentElement
                        || element.parentElement;

                    return parent?.innerText || '';
                }
                """
            )

            score = 0

            if input_type == "search":
                score += 100

            if "搜索" in placeholder:
                score += 80

            if "关键词" in placeholder:
                score += 70

            if "搜索" in aria_label:
                score += 60

            if "热门搜索" in surrounding_text:
                score += 50

            description = (
                f"type={input_type!r}, "
                f"placeholder={placeholder!r}, "
                f"aria-label={aria_label!r}"
            )

            candidates.append((score, candidate, description))

        except Exception:
            continue

    if not candidates:
        raise RuntimeError("没有发现可编辑的文本输入框")

    candidates.sort(key=lambda item: item[0], reverse=True)

    score, candidate, description = candidates[0]

    print(f"使用评分最高的输入框：score={score}, {description}")

    return candidate


async def wait_for_search_input(
    page: Page,
    timeout_seconds: float,
) -> Locator:
    """轮询等待异步渲染的搜索输入框变为可编辑状态。"""

    if timeout_seconds <= 0:
        raise ValueError("--input-timeout 必须大于 0")

    loop = asyncio.get_running_loop()
    deadline = loop.time() + timeout_seconds

    while True:
        try:
            return await find_search_input(page)
        except RuntimeError:
            if loop.time() >= deadline:
                await print_input_candidates(page)
                raise RuntimeError(
                    f"等待 {timeout_seconds:g} 秒后仍未发现可编辑的文本输入框"
                )

            await page.wait_for_timeout(500)


async def submit_search(
    page: Page,
    keyword: str,
    input_timeout: float,
) -> None:
    """填入关键词并触发搜索。"""

    # 记录搜索前表格，用于判断数据是否发生变化。
    table_body = page.locator("tbody").first
    before_table = ""

    try:
        if await table_body.is_visible():
            before_table = await table_body.inner_text()
    except Exception:
        pass

    fill_error: Exception | None = None

    for attempt in range(1, 4):
        search_input = await wait_for_search_input(
            page,
            input_timeout,
        )

        try:
            await search_input.scroll_into_view_if_needed(timeout=5_000)
            await search_input.click(timeout=5_000)
            await search_input.fill(keyword, timeout=5_000)
            fill_error = None
            break
        except PlaywrightTimeoutError as exc:
            fill_error = exc
            print(f"搜索框在点击前被页面替换，正在重试（{attempt}/3）")

    if fill_error is not None:
        raise RuntimeError("多次重试后仍无法填写搜索框") from fill_error

    print(f"已输入关键词：{keyword}")

    clicked = False

    # 优先查找文本为“搜索”或“查询”的按钮。
    button_pattern = re.compile(r"^\s*(搜索|查询)\s*$")

    buttons = page.locator("button").filter(has_text=button_pattern)

    for index in range(await buttons.count()):
        button = buttons.nth(index)

        if await button.is_visible():
            await button.click()
            clicked = True
            print("已点击页面搜索按钮")
            break

    if not clicked:
        # 很多搜索框直接监听 Enter。
        await search_input.press("Enter")
        print("未发现搜索按钮，已按 Enter")

    # 等待表格内容变化。
    if before_table:
        try:
            await page.wait_for_function(
                """
                previousText => {
                    const tbody =
                        document.querySelector('tbody');

                    if (!tbody) {
                        return false;
                    }

                    const currentText =
                        tbody.innerText.trim();

                    return (
                        currentText.length > 0
                        && currentText !== previousText
                    );
                }
                """,
                arg=before_table,
                timeout=30_000,
            )

            print("搜索结果表格已经更新")

        except PlaywrightTimeoutError:
            print("等待表格变化超时，继续等待网络请求完成")

    # SPA 页面可能存在持续连接，不强制等待 networkidle。
    await page.wait_for_timeout(5_000)


async def click_filter(
    page: Page,
    filter_value: str,
    filter_name: str,
) -> None:
    """点击一个结果筛选，并等待表格内容刷新。"""

    options = page.get_by_text(filter_value, exact=True)
    target: Locator | None = None

    for index in range(await options.count()):
        candidate = options.nth(index)

        if await candidate.is_visible():
            target = candidate
            break

    if target is None:
        raise RuntimeError(f"没有发现{filter_name}筛选：{filter_value}")

    table_body = page.locator("tbody").first
    before_table = ""

    try:
        if await table_body.is_visible():
            before_table = await table_body.inner_text()
    except Exception:
        pass

    await target.scroll_into_view_if_needed()
    await target.click()
    print(f"已点击{filter_name}筛选：{filter_value}")

    if before_table:
        try:
            await page.wait_for_function(
                """
                previousText => {
                    const tbody = document.querySelector('tbody');
                    const currentText = tbody?.innerText.trim() || '';
                    return currentText.length > 0 && currentText !== previousText;
                }
                """,
                arg=before_table,
                timeout=30_000,
            )
            print(f"{filter_name}筛选结果表格已经更新")
        except PlaywrightTimeoutError:
            print(f"等待{filter_name}筛选结果变化超时，继续保存当前页面")

    # 给异步请求和页面渲染留出收尾时间。
    await page.wait_for_timeout(2_000)


async def click_result_type(
    page: Page,
    result_type: str,
) -> None:
    """点击结果类型筛选。"""

    await click_filter(
        page=page,
        filter_value=result_type,
        filter_name="类型",
    )


async def click_time_range(
    page: Page,
    time_range: str,
) -> None:
    """点击结果时间范围筛选。"""

    await click_filter(
        page=page,
        filter_value=time_range,
        filter_name="时间",
    )


async def capture_video_detail(
    page: Page,
    row: Locator,
    page_number: int,
    row_number: int,
    record: dict[str, str],
) -> dict[str, str]:
    """点击一条视频的“查看”按钮，从新标签页 URL 中提取 video_id。"""

    title = record["title"]
    view_button = row.get_by_role("button", name="查看", exact=True)
    last_error: Exception | None = None

    for attempt in range(1, 4):
        popup: Page | None = None
        pages_before = set(page.context.pages)

        try:
            async with page.expect_popup(timeout=8_000) as popup_info:
                await view_button.click(
                    timeout=5_000,
                    force=attempt == 3,
                )

            popup = await popup_info.value
            await popup.wait_for_url(
                re.compile(r"/video/detail\?video_id=\d+"),
                timeout=10_000,
            )

            match = re.search(r"[?&]video_id=(\d+)", popup.url)

            if not match:
                raise RuntimeError(f"详情 URL 中没有 video_id：{popup.url}")

            video_id = match.group(1)
            print(f"  [{row_number}] {video_id}  {title}")

            return {
                **record,
                "video_id": video_id,
            }
        except Exception as exc:
            last_error = exc

            # 某些标签页会在 expect_popup 超时边缘才出现，及时关闭。
            delayed_pages = [
                item
                for item in page.context.pages
                if item not in pages_before and item is not page
            ]

            for delayed_page in delayed_pages:
                if not delayed_page.is_closed():
                    await delayed_page.close()

            if attempt < 3:
                print(f"  [{row_number}] 点击未响应，正在重试（{attempt}/3）")
                await page.wait_for_timeout(500)
        finally:
            if popup and not popup.is_closed():
                await popup.close()

    raise RuntimeError("三次点击后仍未打开详情页") from last_error


async def extract_video_list_fields(
    row: Locator,
    title: str,
) -> dict[str, str]:
    """从榜单列表行提取博主与指标字段。"""

    cells = row.locator("td")

    if await cells.count() < 7:
        raise RuntimeError("榜单列表列数不足，无法提取视频指标")

    video_info_lines = [
        line.strip()
        for line in (await cells.nth(1).inner_text()).splitlines()
        if line.strip()
    ]

    fan_line_index = next(
        (
            index
            for index, line in enumerate(video_info_lines)
            if line.startswith("总粉丝数：")
        ),
        None,
    )

    if fan_line_index is None or fan_line_index == 0:
        raise RuntimeError(f"未能从视频信息中提取博主粉丝数：{title}")

    author_name = video_info_lines[fan_line_index - 1]
    total_followers = video_info_lines[fan_line_index].removeprefix("总粉丝数：").strip()

    metrics = [
        (await cells.nth(index).inner_text()).strip()
        for index in range(2, 6)
    ]

    return {
        "author_name": author_name,
        "total_followers": total_followers,
        "hotness": metrics[0],
        "new_views": metrics[1],
        "new_likes": metrics[2],
        "like_rate": metrics[3],
    }


async def extract_video_list_record(
    row: Locator,
    page_number: int,
    row_number: int,
) -> dict[str, str]:
    """提取列表页中可用于去重和写入 Excel 的字段。"""

    title_locator = row.locator('[class*="video-title"]').first

    if not await title_locator.is_visible():
        raise RuntimeError(f"第 {page_number} 页第 {row_number} 行没有发现视频名称")

    title = (await title_locator.inner_text()).strip()
    return {
        "title": title,
        **await extract_video_list_fields(row, title),
    }


async def collect_all_video_details(
    page: Page,
    known_identities: set[tuple[str, str]],
) -> tuple[list[dict[str, str]], int]:
    """采集未出现过的视频详情，并跳过列表页已知视频。"""

    records: list[dict[str, str]] = []
    skipped_count = 0
    page_number = 1

    while True:
        rows = page.locator("tbody tr")
        await rows.first.wait_for(state="visible", timeout=30_000)
        row_count = await rows.count()

        print(f"\n开始采集第 {page_number} 页，共 {row_count} 条视频：")

        for index in range(row_count):
            try:
                list_record = await extract_video_list_record(
                    row=rows.nth(index),
                    page_number=page_number,
                    row_number=index + 1,
                )
                identity = video_identity(
                    list_record["title"],
                    list_record["author_name"],
                )

                if identity in known_identities:
                    skipped_count += 1
                    print(f"  [{index + 1}] 已存在，跳过详情页：{list_record['title']}")
                    continue

                record = await capture_video_detail(
                    page=page,
                    row=rows.nth(index),
                    page_number=page_number,
                    row_number=index + 1,
                    record=list_record,
                )
                records.append(record)
                known_identities.add(identity)
            except Exception as exc:
                print(
                    f"  [{index + 1}] 获取 video_id 失败：{exc}",
                    file=sys.stderr,
                )

        next_button = page.locator(
            ".arco-pagination-item-next, .arco-pagination-next"
        ).first

        if not await next_button.is_visible():
            break

        next_class = await next_button.get_attribute("class") or ""
        aria_disabled = await next_button.get_attribute("aria-disabled")

        if "disabled" in next_class or aria_disabled == "true":
            break

        first_row_text = await rows.first.inner_text()
        await next_button.click()

        try:
            await page.wait_for_function(
                """
                previousText => {
                    const firstRow = document.querySelector('tbody tr');
                    return firstRow && firstRow.innerText !== previousText;
                }
                """,
                arg=first_row_text,
                timeout=30_000,
            )
        except PlaywrightTimeoutError:
            print("等待下一页表格更新超时，停止翻页", file=sys.stderr)
            break

        page_number += 1
        await page.wait_for_timeout(1_000)

    print(f"\n共获取 {len(records)} 个新 video_id，跳过 {skipped_count} 条已有视频")
    return records, skipped_count


def excel_sheet_name(keyword: str) -> str:
    """将关键词转换为 Excel 允许的工作表名称。"""

    sheet_name = re.sub(r'[:\\/?*\[\]]', "_", keyword).strip()

    if not sheet_name:
        raise ValueError("关键词无法转换为有效的 Excel 工作表名称")

    return sheet_name[:31]


def video_identity(title: object, author_name: object) -> tuple[str, str]:
    """使用视频名称与博主名称标识一条视频。"""

    return str(title).strip(), str(author_name).strip()


def worksheet_video_identities(worksheet) -> set[tuple[str, str]]:
    """读取工作表中已有的“视频名称 + 博主名称”组合。"""

    headers = [cell.value for cell in worksheet[1]]

    try:
        title_index = headers.index("视频名称")
        author_index = headers.index("博主名称")
    except ValueError:
        return set()

    return {
        video_identity(row[title_index], row[author_index])
        for row in worksheet.iter_rows(min_row=2, values_only=True)
        if len(row) > max(title_index, author_index)
        and row[title_index]
        and row[author_index]
    }


def existing_video_identities(
    excel_path: Path,
    keyword: str,
) -> set[tuple[str, str]]:
    """读取指定关键词 Sheet 已保存的视频，用于跳过详情页。"""

    if not excel_path.exists():
        return set()

    workbook = load_workbook(excel_path, read_only=True, data_only=True)

    try:
        sheet_name = excel_sheet_name(keyword)

        if sheet_name not in workbook.sheetnames:
            return set()

        return worksheet_video_identities(workbook[sheet_name])
    finally:
        workbook.close()


def write_result_excel(
    records: list[dict[str, str | int]],
    excel_path: Path,
    keyword: str,
    result_type: str,
    time_range: str,
    crawled_at: str,
) -> tuple[int, int]:
    """按“视频名称 + 博主名称”去重，增量更新工作表。"""

    excel_path.parent.mkdir(parents=True, exist_ok=True)
    sheet_name = excel_sheet_name(keyword)

    if excel_path.exists():
        workbook = load_workbook(excel_path)

        if sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
        else:
            worksheet = workbook.create_sheet(sheet_name)
    else:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = sheet_name

    existing_headers = [
        cell.value.strip() if isinstance(cell.value, str) else cell.value
        for cell in worksheet[1]
    ]

    if not any(existing_headers):
        worksheet.delete_rows(1, worksheet.max_row)
        worksheet.append(RESULT_HEADERS)
    elif existing_headers != RESULT_HEADERS:
        old_rows = list(worksheet.iter_rows(min_row=2, values_only=True))
        old_header_positions = {
            header: index
            for index, header in enumerate(existing_headers)
            if header
        }

        worksheet.delete_rows(1, worksheet.max_row)
        worksheet.append(RESULT_HEADERS)

        for old_row in old_rows:
            worksheet.append(
                [
                    (
                        old_row[old_header_positions[header]]
                        if header in old_header_positions
                        and old_header_positions[header] < len(old_row)
                        else ""
                    )
                    for header in RESULT_HEADERS
                ]
            )

    for cell in worksheet[1]:
        cell.font = Font(bold=True)

    known_identities = worksheet_video_identities(worksheet)
    existing_numbers = [
        int(row[0])
        for row in worksheet.iter_rows(min_row=2, values_only=True)
        if len(row) >= 1
        and isinstance(row[0], (int, float))
        and float(row[0]).is_integer()
    ]
    next_number = max(existing_numbers, default=0) + 1
    added_count = 0
    skipped_count = 0

    for record in records:
        video_url = f"{DOUYIN_VIDEO_URL_PREFIX}{record['video_id']}"
        identity = video_identity(record["title"], record["author_name"])

        if identity in known_identities:
            skipped_count += 1
            continue

        worksheet.append(
            [
                next_number,
                result_type,
                crawled_at,
                time_range,
                record["title"],
                video_url,
                record["author_name"],
                record["total_followers"],
                record["hotness"],
                record["new_views"],
                record["new_likes"],
                record["like_rate"],
            ]
        )

        url_cell = worksheet.cell(row=worksheet.max_row, column=6)
        url_cell.hyperlink = str(url_cell.value)
        url_cell.style = "Hyperlink"
        known_identities.add(identity)
        next_number += 1
        added_count += 1

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    worksheet.column_dimensions["A"].width = 8
    worksheet.column_dimensions["B"].width = 14
    worksheet.column_dimensions["C"].width = 21
    worksheet.column_dimensions["D"].width = 14
    worksheet.column_dimensions["E"].width = 70
    worksheet.column_dimensions["F"].width = 48
    worksheet.column_dimensions["G"].width = 24
    worksheet.column_dimensions["H"].width = 16
    worksheet.column_dimensions["I"].width = 14
    worksheet.column_dimensions["J"].width = 16
    worksheet.column_dimensions["K"].width = 16
    worksheet.column_dimensions["L"].width = 12

    workbook.save(excel_path)
    return added_count, skipped_count


async def main() -> None:
    args = parse_args()

    if not PROFILE_PATH.exists():
        raise FileNotFoundError(
            f"没有找到 douhot Profile：{PROFILE_PATH}\n"
            "请先运行 `uv run crwl profiles`，"
            "并确认已经按 q 保存。"
        )

    excel_path = RESULT_EXCEL_PATH.resolve()

    captured_videos: list[dict[str, str | int]] = []
    known_video_identities = existing_video_identities(
        excel_path,
        args.keyword,
    )
    skipped_in_list = 0

    if known_video_identities:
        print(f"已加载 {len(known_video_identities)} 条已有视频，将跳过详情页")

    browser_config = BrowserConfig(
        browser_type="chromium",
        headless=args.headless,
        use_managed_browser=True,
        use_persistent_context=True,
        user_data_dir=str(PROFILE_PATH),
        viewport_width=1440,
        viewport_height=1000,
        verbose=True,
    )

    run_config = CrawlerRunConfig(
        cache_mode=CacheMode.BYPASS,
        page_timeout=120_000,
        delay_before_return_html=2.0,
    )

    crawler = AsyncWebCrawler(config=browser_config)

    async def after_goto(
        page: Page,
        context,
        url: str,
        response,
        **kwargs,
    ) -> Page:
        nonlocal skipped_in_list

        print(f"\n页面已打开：{url}")

        await page.wait_for_load_state("domcontentloaded")

        # 搜索框由 React/Vue 异步渲染，等待它真正变为可编辑状态。
        await wait_for_search_input(
            page,
            args.input_timeout,
        )

        await print_input_candidates(page)

        await submit_search(
            page=page,
            keyword=args.keyword,
            input_timeout=args.input_timeout,
        )

        await click_result_type(
            page=page,
            result_type=args.result_type,
        )

        await click_time_range(
            page=page,
            time_range=args.time_range,
        )

        new_records, skipped_count = await collect_all_video_details(
            page,
            known_video_identities,
        )
        captured_videos.extend(new_records)
        skipped_in_list += skipped_count

        return page

    crawler.crawler_strategy.set_hook(
        "after_goto",
        after_goto,
    )

    try:
        await crawler.start()

        result = await crawler.arun(
            url=TARGET_URL,
            config=run_config,
        )

        if not result.success:
            raise RuntimeError(result.error_message or "爬取失败")

        crawled_at = datetime.now().astimezone().strftime("%Y-%m-%d %H:%M:%S")
        added_count, skipped_count = write_result_excel(
            records=captured_videos,
            excel_path=excel_path,
            keyword=args.keyword,
            result_type=args.result_type,
            time_range=args.time_range,
            crawled_at=crawled_at,
        )

        print(f"\n总结果 Excel：{excel_path}（Sheet：{excel_sheet_name(args.keyword)}）")
        print(
            f"本次新增 {added_count} 条，"
            f"跳过已有视频 {skipped_in_list + skipped_count} 条"
        )

    finally:
        await crawler.close()


if __name__ == "__main__":
    try:
        asyncio.run(main())
    except KeyboardInterrupt:
        raise SystemExit(130)
    except Exception as exc:
        print(f"\n错误：{exc}", file=sys.stderr)
        raise SystemExit(1)
