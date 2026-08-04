from __future__ import annotations

import asyncio
import hashlib
import hmac
import json
import os
import re
import shutil
import sqlite3
import threading
import time
import uuid
from collections.abc import Awaitable, Callable
from dataclasses import asdict, dataclass
from datetime import UTC, datetime, timedelta
from pathlib import Path
from typing import Any
from urllib.parse import quote

from openpyxl import load_workbook

from douhot_crawler.browser.cookies import inspect_douhot_cookie
from douhot_crawler.browser.setup import chromium_status, detect_system_browser
from douhot_crawler.core.config import (
    EXTRACT_API_URL,
    LOGIN_URL,
    RESULT_EXCEL_PATH,
    RESULT_TYPE_CHOICES,
    TIME_RANGE_CHOICES,
)
from douhot_crawler.core.models import RunOptions
from douhot_crawler.core.storage import write_result_excel
from douhot_crawler.crawling.runner import run as run_crawl
from douhot_crawler.transcript.analyzer import (
    analyze_excel,
    extract_transcript,
    read_cookie,
)
from douhot_crawler.transcript.cookies import inspect_transcript_cookie

ACTIVE_STATUSES = {"queued", "running", "waiting_login"}
TERMINAL_STATUSES = {"succeeded", "failed", "cancelled"}
VIDEO_URL_HEADER = "视频的url"
VIDEO_ID_PATTERN = re.compile(r"/video/(\d+)")
KEYWORD_CACHE_TTL = timedelta(hours=4)


def utc_now() -> str:
    return datetime.now(UTC).isoformat()


def user_key(user_id: str) -> str:
    normalized = user_id.strip()
    if (
        not normalized
        or len(normalized) > 255
        or any(ord(ch) < 32 for ch in normalized)
    ):
        raise ValueError("user_id 必须是 1-255 个可打印字符")
    return hashlib.sha256(normalized.encode("utf-8")).hexdigest()


@dataclass(frozen=True, slots=True)
class ServiceSettings:
    data_root: Path
    public_url: str
    download_secret: str
    login_timeout_seconds: int = 300
    cookie_source: Path | None = None

    @classmethod
    def from_env(cls) -> ServiceSettings:
        default_root = RESULT_EXCEL_PATH.parent.parent
        default_cookie_source = (
            Path(__file__).resolve().parent.parent.parent / "cookie.config"
        )
        return cls(
            data_root=Path(os.environ.get("DOUHOT_DATA_ROOT", default_root))
            .expanduser()
            .resolve(),
            public_url=os.environ.get(
                "DOUHOT_PUBLIC_URL", "http://127.0.0.1:8765"
            ).rstrip("/"),
            download_secret=os.environ.get(
                "DOUHOT_DOWNLOAD_SECRET", "change-me-in-production"
            ),
            login_timeout_seconds=int(
                os.environ.get("DOUHOT_LOGIN_TIMEOUT_SECONDS", "300")
            ),
            cookie_source=Path(
                os.environ.get("DOUHOT_COOKIE_SOURCE", default_cookie_source)
            )
            .expanduser()
            .resolve(),
        )


@dataclass(frozen=True, slots=True)
class UserPaths:
    root: Path
    profile: Path
    cookie: Path
    jobs: Path


class JobStore:
    def __init__(self, path: Path) -> None:
        path.parent.mkdir(parents=True, exist_ok=True)
        self.path = path
        self._lock = threading.RLock()
        self._initialize()

    def _connect(self) -> sqlite3.Connection:
        connection = sqlite3.connect(self.path)
        connection.row_factory = sqlite3.Row
        return connection

    def _initialize(self) -> None:
        with self._connect() as connection:
            connection.execute(
                """
                CREATE TABLE IF NOT EXISTS jobs (
                    id TEXT PRIMARY KEY,
                    user_key TEXT NOT NULL,
                    kind TEXT NOT NULL,
                    status TEXT NOT NULL,
                    params_json TEXT NOT NULL,
                    result_path TEXT,
                    result_mime TEXT,
                    result_json TEXT,
                    error TEXT,
                    created_at TEXT NOT NULL,
                    updated_at TEXT NOT NULL
                )
                """
            )
            connection.execute(
                "UPDATE jobs SET status = 'failed', error = ?, updated_at = ? "
                "WHERE status IN ('queued', 'running', 'waiting_login')",
                ("服务重启，未完成任务已终止", utc_now()),
            )

    def create(
        self, *, owner: str, kind: str, params: dict[str, Any]
    ) -> dict[str, Any]:
        job_id = str(uuid.uuid4())
        timestamp = utc_now()
        with self._lock, self._connect() as connection:
            active = connection.execute(
                "SELECT id FROM jobs WHERE user_key = ? AND status IN "
                "('queued', 'running', 'waiting_login')",
                (owner,),
            ).fetchone()
            if active is not None:
                raise ValueError(f"当前用户已有运行中的任务：{active['id']}")
            connection.execute(
                "INSERT INTO jobs VALUES (?, ?, ?, 'queued', ?, NULL, NULL, NULL, NULL, ?, ?)",
                (
                    job_id,
                    owner,
                    kind,
                    json.dumps(params, ensure_ascii=False),
                    timestamp,
                    timestamp,
                ),
            )
        return self.get(job_id, owner)

    def update(self, job_id: str, owner: str, **values: Any) -> dict[str, Any]:
        allowed = {"status", "result_path", "result_mime", "result_json", "error"}
        unknown = set(values) - allowed
        if unknown:
            raise ValueError(f"不支持更新任务字段：{sorted(unknown)}")
        values["updated_at"] = utc_now()
        assignments = ", ".join(f"{name} = ?" for name in values)
        serialized = [
            json.dumps(value, ensure_ascii=False)
            if name == "result_json" and value is not None
            else value
            for name, value in values.items()
        ]
        with self._lock, self._connect() as connection:
            cursor = connection.execute(
                f"UPDATE jobs SET {assignments} WHERE id = ? AND user_key = ?",
                (*serialized, job_id, owner),
            )
            if cursor.rowcount != 1:
                raise ValueError("任务不存在")
        return self.get(job_id, owner)

    def get(self, job_id: str, owner: str) -> dict[str, Any]:
        with self._connect() as connection:
            row = connection.execute(
                "SELECT * FROM jobs WHERE id = ? AND user_key = ?", (job_id, owner)
            ).fetchone()
            if row is None and re.fullmatch(r"[0-9a-fA-F-]{8,35}", job_id):
                matches = connection.execute(
                    "SELECT * FROM jobs WHERE user_key = ? AND id LIKE ? "
                    "ORDER BY created_at DESC LIMIT 2",
                    (owner, f"{job_id}%"),
                ).fetchall()
                if len(matches) > 1:
                    raise ValueError("任务 ID 前缀不唯一，请使用完整任务 ID")
                row = matches[0] if matches else None
        if row is None:
            raise ValueError("任务不存在")
        result = dict(row)
        result["params"] = json.loads(result.pop("params_json"))
        raw_result = result.pop("result_json")
        result["result"] = json.loads(raw_result) if raw_result else None
        return result

    def active(self, owner: str) -> dict[str, Any] | None:
        with self._lock, self._connect() as connection:
            row = connection.execute(
                "SELECT id FROM jobs WHERE user_key = ? AND status IN "
                "('queued', 'running', 'waiting_login') "
                "ORDER BY created_at DESC LIMIT 1",
                (owner,),
            ).fetchone()
        return self.get(row["id"], owner) if row is not None else None

    def latest(self, owner: str, *, kinds: tuple[str, ...]) -> dict[str, Any]:
        if not kinds:
            raise ValueError("必须指定任务类型")
        placeholders = ", ".join("?" for _ in kinds)
        with self._lock, self._connect() as connection:
            row = connection.execute(
                f"SELECT id FROM jobs WHERE user_key = ? AND kind IN ({placeholders}) "
                "ORDER BY created_at DESC LIMIT 1",
                (owner, *kinds),
            ).fetchone()
        if row is None:
            raise ValueError("任务不存在")
        return self.get(row["id"], owner)

    def latest_crawl_by_keyword(
        self, owner: str, keyword: str
    ) -> dict[str, Any] | None:
        normalized = " ".join(keyword.split()).casefold()
        if not normalized:
            raise ValueError("keyword 不能为空")
        cutoff = (datetime.now(UTC) - KEYWORD_CACHE_TTL).isoformat()
        with self._lock, self._connect() as connection:
            rows = connection.execute(
                "SELECT id, params_json FROM jobs "
                "WHERE user_key = ? AND kind = 'crawl' AND status = 'succeeded' "
                "AND updated_at > ? "
                "ORDER BY created_at DESC",
                (owner, cutoff),
            ).fetchall()
        for row in rows:
            params = json.loads(row["params_json"])
            candidate = " ".join(str(params.get("keyword") or "").split()).casefold()
            if candidate == normalized:
                return self.get(row["id"], owner)
        return None


Runner = Callable[[str, str, asyncio.Event], Awaitable[dict[str, Any]]]
StatusCallback = Callable[[dict[str, Any]], Awaitable[None]]


class JobManager:
    def __init__(self, settings: ServiceSettings | None = None) -> None:
        self.settings = settings or ServiceSettings.from_env()
        self.settings.data_root.mkdir(parents=True, exist_ok=True)
        self.store = JobStore(self.settings.data_root / "jobs.sqlite3")
        self._tasks: dict[str, asyncio.Task[None]] = {}
        self._cancellations: dict[str, asyncio.Event] = {}

    def request_shutdown(self) -> None:
        """请求所有后台任务在当前记录落盘后停止。"""

        for cancellation in tuple(self._cancellations.values()):
            cancellation.set()

    async def shutdown(self, *, timeout: float = 30.0) -> None:
        """停止并等待后台任务，超时后再取消仍未结束的任务。"""

        self.request_shutdown()
        tasks = tuple(self._tasks.values())
        if not tasks:
            return
        _, pending = await asyncio.wait(tasks, timeout=timeout)
        for task in pending:
            task.cancel()
        if pending:
            await asyncio.gather(*pending, return_exceptions=True)

    async def wait_for_terminal(
        self,
        user_id: str,
        job_id: str | None = None,
        *,
        kinds: tuple[str, ...] = ("crawl", "analyze"),
        timeout: float = 900.0,
        poll_interval: float = 1.0,
        on_status: StatusCallback | None = None,
    ) -> dict[str, Any]:
        """等待后台任务进入终态，并允许 MCP 请求期间持续报告进度。"""

        if timeout <= 0:
            raise ValueError("timeout 必须大于 0")
        if poll_interval <= 0:
            raise ValueError("poll_interval 必须大于 0")

        job = self.describe(user_id, job_id, kinds=kinds)
        deadline = asyncio.get_running_loop().time() + timeout
        while job["status"] in ACTIVE_STATUSES:
            if on_status is not None:
                await on_status(job)
            remaining = deadline - asyncio.get_running_loop().time()
            if remaining <= 0:
                raise TimeoutError(
                    f"等待任务 {job['id']} 完成超过 {timeout:g} 秒；任务仍在后台运行"
                )
            task = self._tasks.get(job["id"])
            interval = min(poll_interval, remaining)
            if task is None:
                await asyncio.sleep(interval)
            else:
                await asyncio.wait({task}, timeout=interval)
            job = self.describe(user_id, job["id"], kinds=kinds)

        if on_status is not None:
            await on_status(job)
        return job

    def paths(self, user_id: str) -> UserPaths:
        root = self.settings.data_root / "users" / user_key(user_id)
        paths = UserPaths(
            root=root,
            profile=root / "profile",
            cookie=root / "cookie.config",
            jobs=root / "jobs",
        )
        paths.jobs.mkdir(parents=True, exist_ok=True)
        if (
            self.settings.cookie_source is not None
            and self.settings.cookie_source.is_file()
        ):
            shutil.copyfile(self.settings.cookie_source, paths.cookie)
            paths.cookie.chmod(0o600)
        return paths

    async def _start(
        self,
        *,
        user_id: str,
        kind: str,
        params: dict[str, Any],
        runner: Runner,
    ) -> dict[str, Any]:
        owner = user_key(user_id)
        job = self.store.create(owner=owner, kind=kind, params=params)
        cancellation = asyncio.Event()
        self._cancellations[job["id"]] = cancellation

        async def execute() -> None:
            try:
                self.store.update(job["id"], owner, status="running")
                result = await runner(job["id"], owner, cancellation)
                if cancellation.is_set():
                    self.store.update(job["id"], owner, status="cancelled")
                else:
                    self.store.update(
                        job["id"],
                        owner,
                        status="succeeded",
                        result_path=result.pop("result_path", None),
                        result_mime=result.pop("result_mime", None),
                        result_json=result,
                    )
            except asyncio.CancelledError:
                self.store.update(job["id"], owner, status="cancelled")
            except Exception as exc:  # noqa: BLE001 - job failures must reach the store
                self.store.update(job["id"], owner, status="failed", error=str(exc))
            finally:
                self._cancellations.pop(job["id"], None)
                self._tasks.pop(job["id"], None)

        self._tasks[job["id"]] = asyncio.create_task(execute())
        return self.describe(user_id, job["id"])

    async def start_crawl(
        self,
        user_id: str,
        *,
        keyword: str,
        result_type: str,
        time_range: str,
        input_timeout: float,
        detail_delay: float,
        limit: int = 10,
    ) -> dict[str, Any]:
        if not keyword.strip():
            raise ValueError("keyword 不能为空")
        if result_type not in RESULT_TYPE_CHOICES:
            raise ValueError(
                f"result_type 必须是：{', '.join(RESULT_TYPE_CHOICES)}"
            )
        if time_range not in TIME_RANGE_CHOICES:
            raise ValueError(f"time_range 必须是：{', '.join(TIME_RANGE_CHOICES)}")
        if limit < 1:
            raise ValueError("limit 必须大于 0")
        browser_ok, browser_detail = chromium_status()
        if not browser_ok:
            raise ValueError(browser_detail)
        paths = self.paths(user_id)
        options = RunOptions(
            keyword=keyword.strip(),
            input_timeout=input_timeout,
            detail_delay=detail_delay,
            result_type=result_type,
            time_range=time_range,
            headless=True,
        )

        async def runner(
            job_id: str, _owner: str, cancellation: asyncio.Event
        ) -> dict[str, Any]:
            output = paths.jobs / job_id / "result.xlsx"
            output.parent.mkdir(parents=True, exist_ok=True)
            result = await run_crawl(
                options,
                stop_requested=cancellation.is_set,
                profile_path=paths.profile,
                excel_path=output,
                max_results=limit,
            )
            if not output.is_file():
                await asyncio.to_thread(
                    write_result_excel,
                    [],
                    output,
                    options.keyword,
                    options.result_type,
                    options.time_range,
                    utc_now(),
                )
            return {
                **result,
                "row_count": self._row_count(output),
                "sha256": self._sha256(output),
                "result_path": str(output),
                "result_mime": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            }

        return await self._start(
            user_id=user_id,
            kind="crawl",
            params={
                "keyword": keyword,
                "result_type": result_type,
                "time_range": time_range,
                "limit": limit,
            },
            runner=runner,
        )

    async def start_analyze(
        self,
        user_id: str,
        *,
        crawl_job_id: str | None = None,
        overwrite: bool = False,
        limit: int | None = None,
    ) -> dict[str, Any]:
        owner = user_key(user_id)
        source = (
            self.store.get(crawl_job_id, owner)
            if crawl_job_id
            else self.store.latest(owner, kinds=("crawl",))
        )
        if source["status"] != "succeeded" or not source["result_path"]:
            raise ValueError("爬取任务尚未生成可分析的 Excel")
        source_path = Path(source["result_path"]).resolve()
        paths = self.paths(user_id)

        async def runner(
            job_id: str, _owner: str, cancellation: asyncio.Event
        ) -> dict[str, Any]:
            output = paths.jobs / job_id / "result.xlsx"
            output.parent.mkdir(parents=True, exist_ok=True)
            await asyncio.to_thread(output.write_bytes, source_path.read_bytes())
            args = type(
                "AnalyzeArgs",
                (),
                {
                    "excel": output,
                    "cookie_file": paths.cookie,
                    "sheet": None,
                    "callback_url": "",
                    "timeout": 90.0,
                    "delay": 0.0,
                    "limit": limit,
                    "overwrite": overwrite,
                },
            )()
            counts = await asyncio.to_thread(
                analyze_excel, args, stop_requested=cancellation.is_set
            )
            return {
                "success_count": counts[0],
                "skipped_count": counts[1],
                "failed_count": counts[2],
                "row_count": self._row_count(output),
                "sha256": self._sha256(output),
                "result_path": str(output),
                "result_mime": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            }

        return await self._start(
            user_id=user_id,
            kind="analyze",
            params={
                "crawl_job_id": source["id"],
                "overwrite": overwrite,
                "limit": limit,
            },
            runner=runner,
        )

    async def start_login(self, user_id: str) -> dict[str, Any]:
        owner = user_key(user_id)
        active = self.store.active(owner)
        if active is not None:
            if active["kind"] == "login":
                return self.describe(user_id, active["id"])
            raise ValueError(f"当前用户已有运行中的任务：{active['id']}")
        browser_ok, browser_detail = chromium_status()
        if not browser_ok:
            raise ValueError(browser_detail)
        paths = self.paths(user_id)

        async def runner(
            job_id: str, owner: str, cancellation: asyncio.Event
        ) -> dict[str, Any]:
            from playwright.async_api import async_playwright

            qr_path = paths.jobs / job_id / "qr.png"
            qr_path.parent.mkdir(parents=True, exist_ok=True)
            paths.profile.mkdir(parents=True, exist_ok=True)
            async with async_playwright() as playwright:
                channel, _ = detect_system_browser()
                kwargs: dict[str, Any] = {
                    "user_data_dir": str(paths.profile),
                    "headless": True,
                    "viewport": {"width": 1440, "height": 1000},
                }
                if channel:
                    kwargs["channel"] = channel
                context = await playwright.chromium.launch_persistent_context(**kwargs)
                try:
                    page = (
                        context.pages[0] if context.pages else await context.new_page()
                    )
                    await page.goto(
                        LOGIN_URL, wait_until="domcontentloaded", timeout=120_000
                    )
                    await page.wait_for_timeout(2_000)
                    await self._capture_qr(page, qr_path)
                    self.store.update(
                        job_id,
                        owner,
                        status="waiting_login",
                        result_path=str(qr_path),
                        result_mime="image/png",
                        result_json={"message": "请使用抖音扫描二维码"},
                    )
                    deadline = (
                        asyncio.get_running_loop().time()
                        + self.settings.login_timeout_seconds
                    )
                    while not cancellation.is_set():
                        cookies = await context.cookies()
                        names = {cookie["name"] for cookie in cookies}
                        if names & {
                            "sessionid_douhot",
                            "sessionid_ss_douhot",
                            "sid_tt_douhot",
                        }:
                            return {
                                "message": "扫码登录成功，Profile 已保存",
                                "result_path": str(qr_path),
                                "result_mime": "image/png",
                            }
                        if asyncio.get_running_loop().time() >= deadline:
                            raise TimeoutError("扫码登录超时，请重新发起")
                        await asyncio.sleep(1)
                    return {"message": "登录已取消"}
                finally:
                    await context.close()

        return await self._start(
            user_id=user_id, kind="login", params={}, runner=runner
        )

    async def cancel(
        self,
        user_id: str,
        job_id: str | None,
        *,
        kinds: tuple[str, ...],
    ) -> dict[str, Any]:
        owner = user_key(user_id)
        job = (
            self.store.get(job_id, owner)
            if job_id
            else self.store.latest(owner, kinds=kinds)
        )
        canonical_id = job["id"]
        if job["status"] in TERMINAL_STATUSES:
            return self.describe(user_id, canonical_id)
        cancellation = self._cancellations.get(canonical_id)
        if cancellation is not None:
            cancellation.set()
        else:
            self.store.update(canonical_id, owner, status="cancelled")
        return self.describe(user_id, canonical_id)

    def describe(
        self,
        user_id: str,
        job_id: str | None,
        *,
        kinds: tuple[str, ...] = ("login", "crawl", "analyze"),
    ) -> dict[str, Any]:
        owner = user_key(user_id)
        job = (
            self.store.get(job_id, owner)
            if job_id
            else self.store.latest(owner, kinds=kinds)
        )
        path = Path(job["result_path"]) if job["result_path"] else None
        if path is not None and path.is_file():
            job["download_url"] = self.signed_download_url(user_id, job["id"])
            job["size_bytes"] = path.stat().st_size
        return job

    def list_videos(
        self,
        user_id: str,
        job_id: str | None,
        *,
        keyword: str | None = None,
        offset: int,
        limit: int,
    ) -> dict[str, Any]:
        if offset < 0 or limit < 1 or limit > 100:
            raise ValueError("offset 必须不小于 0，limit 必须在 1-100")
        owner = user_key(user_id)
        if job_id:
            job = self.store.get(job_id, owner)
        elif keyword is not None:
            normalized_keyword = " ".join(keyword.split())
            job = self.store.latest_crawl_by_keyword(owner, normalized_keyword)
            if job is None:
                return {
                    "found": False,
                    "keyword": normalized_keyword,
                    "total": 0,
                    "offset": offset,
                    "limit": limit,
                    "videos": [],
                }
        else:
            job = self.store.latest(owner, kinds=("crawl",))
        if job["status"] != "succeeded" or not job["result_path"]:
            raise ValueError("任务尚未生成候选视频")
        workbook = load_workbook(job["result_path"], read_only=True, data_only=True)
        rows: list[dict[str, Any]] = []
        try:
            for sheet in workbook.worksheets:
                iterator = sheet.iter_rows(values_only=True)
                headers = [str(value or "") for value in next(iterator, ())]
                if VIDEO_URL_HEADER not in headers:
                    continue
                for values in iterator:
                    record = dict(zip(headers, values, strict=False))
                    url = str(record.get(VIDEO_URL_HEADER) or "")
                    match = VIDEO_ID_PATTERN.search(url)
                    video_id = (
                        match.group(1)
                        if match
                        else hashlib.sha256(url.encode()).hexdigest()[:16]
                    )
                    rows.append(
                        {
                            "index": len(rows) + 1,
                            "video_id": video_id,
                            "title": record.get("视频名称"),
                            "author": record.get("博主名称"),
                            "heat": record.get("热度值"),
                            "url": url,
                            "sheet": sheet.title,
                        }
                    )
        finally:
            workbook.close()
        return {
            "found": True,
            "job_id": job["id"],
            "keyword": job["params"].get("keyword"),
            "total": len(rows),
            "offset": offset,
            "limit": limit,
            "videos": rows[offset : offset + limit],
        }

    async def transcript(self, user_id: str, share_link: str) -> dict[str, Any]:
        if not share_link.strip():
            raise ValueError("share_link 不能为空")
        cookie = await asyncio.to_thread(read_cookie, self.paths(user_id).cookie)
        transcript = await asyncio.to_thread(
            extract_transcript, share_link.strip(), cookie, "", 90.0
        )
        return {"share_link": share_link.strip(), "transcript": transcript}

    def health(self, user_id: str) -> dict[str, Any]:
        paths = self.paths(user_id)
        browser_ok, browser_detail = chromium_status()
        login = inspect_douhot_cookie(paths.profile)
        transcript = inspect_transcript_cookie(paths.cookie)
        return {
            "ok": browser_ok,
            "browser": {"ok": browser_ok, "detail": browser_detail},
            "douhot_login": asdict(login),
            "transcript_cookie": asdict(transcript),
            "extract_api_configured": bool(EXTRACT_API_URL),
        }

    def signed_download_url(self, user_id: str, job_id: str) -> str:
        owner = user_key(user_id)
        expires = int(time.time()) + 900
        signature = hmac.new(
            self.settings.download_secret.encode(),
            f"{job_id}:{owner}:{expires}".encode(),
            hashlib.sha256,
        ).hexdigest()
        return (
            f"{self.settings.public_url}/downloads/{quote(job_id)}"
            f"?owner={owner}&expires={expires}&signature={signature}"
        )

    def resolve_download(
        self, job_id: str, owner: str, expires: str, signature: str
    ) -> tuple[Path, str]:
        try:
            expiry = int(expires)
        except ValueError as exc:
            raise ValueError("下载地址已失效") from exc
        if expiry < int(time.time()):
            raise ValueError("下载地址已过期")
        expected = hmac.new(
            self.settings.download_secret.encode(),
            f"{job_id}:{owner}:{expiry}".encode(),
            hashlib.sha256,
        ).hexdigest()
        if not hmac.compare_digest(signature, expected):
            raise ValueError("下载签名无效")
        job = self.store.get(job_id, owner)
        if not job["result_path"]:
            raise ValueError("任务没有可下载产物")
        path = Path(job["result_path"]).resolve()
        allowed = (self.settings.data_root / "users" / owner / "jobs").resolve()
        if not path.is_relative_to(allowed) or not path.is_file():
            raise ValueError("任务产物不存在")
        return path, job["result_mime"] or "application/octet-stream"

    @staticmethod
    async def _capture_qr(page, output: Path, *, timeout_ms: int = 30_000) -> None:
        selectors = (
            "img[src*='qr']",
            "[class*='qrcode'] img",
            "[class*='qr-code'] img",
            "img",
            "canvas",
            "svg[class*='qrcode']",
            "svg[class*='qr-code']",
        )
        deadline = asyncio.get_running_loop().time() + timeout_ms / 1_000
        while True:
            for selector in selectors:
                candidates = page.locator(selector)
                for index in range(min(await candidates.count(), 10)):
                    candidate = candidates.nth(index)
                    if not await candidate.is_visible():
                        continue
                    box = await candidate.bounding_box()
                    if not box:
                        continue
                    width = box["width"]
                    height = box["height"]
                    if min(width, height) >= 150 and 0.8 <= width / height <= 1.25:
                        await candidate.screenshot(path=str(output))
                        return
            if asyncio.get_running_loop().time() >= deadline:
                raise TimeoutError("登录页未出现可扫描二维码，请重新发起登录")
            await page.wait_for_timeout(500)

    @staticmethod
    def _sha256(path: Path) -> str:
        digest = hashlib.sha256()
        with path.open("rb") as handle:
            for chunk in iter(lambda: handle.read(1024 * 1024), b""):
                digest.update(chunk)
        return digest.hexdigest()

    @staticmethod
    def _row_count(path: Path) -> int:
        workbook = load_workbook(path, read_only=True)
        try:
            return sum(max(sheet.max_row - 1, 0) for sheet in workbook.worksheets)
        finally:
            workbook.close()
