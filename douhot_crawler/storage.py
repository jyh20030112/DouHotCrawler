"""增量 Excel 结果库。"""

import re
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

from .config import DOUYIN_VIDEO_URL_PREFIX, RESULT_HEADERS
from .models import VideoIdentity, VideoRecord, video_identity


def excel_sheet_name(keyword: str) -> str:
    """将关键词转换为 Excel 允许的工作表名称。"""

    sheet_name = re.sub(r'[:\\/?*\[\]]', "_", keyword).strip()
    if not sheet_name:
        raise ValueError("关键词无法转换为有效的 Excel 工作表名称")
    return sheet_name[:31]


def worksheet_video_identities(worksheet) -> set[VideoIdentity]:
    """读取工作表中已有的“视频名称 + 博主名称”组合。"""

    headers = [cell.value for cell in worksheet[1]]
    try:
        title_index = headers.index("视频名称")
        author_index = headers.index("博主名称")
    except ValueError:
        return set()

    return {
        video_identity(row[title_index], row[author_index])
        for row in worksheet.iter_rows(min_row=2, values_only=True)
        if len(row) > max(title_index, author_index)
        and row[title_index]
        and row[author_index]
    }


def existing_video_identities(
    excel_path: Path,
    keyword: str,
) -> set[VideoIdentity]:
    """读取指定关键词 Sheet 已保存的视频，用于跳过详情页。"""

    if not excel_path.exists():
        return set()

    workbook = load_workbook(excel_path, read_only=True, data_only=True)
    try:
        sheet_name = excel_sheet_name(keyword)
        if sheet_name not in workbook.sheetnames:
            return set()
        return worksheet_video_identities(workbook[sheet_name])
    finally:
        workbook.close()


def ensure_result_headers(worksheet) -> None:
    """创建或迁移工作表表头，同时保留既有记录。"""

    existing_headers = [
        cell.value.strip() if isinstance(cell.value, str) else cell.value
        for cell in worksheet[1]
    ]
    if not any(existing_headers):
        worksheet.delete_rows(1, worksheet.max_row)
        worksheet.append(RESULT_HEADERS)
    elif existing_headers != RESULT_HEADERS:
        old_rows = list(worksheet.iter_rows(min_row=2, values_only=True))
        old_header_positions = {
            header: index
            for index, header in enumerate(existing_headers)
            if header
        }
        worksheet.delete_rows(1, worksheet.max_row)
        worksheet.append(RESULT_HEADERS)
        for old_row in old_rows:
            worksheet.append(
                [
                    (
                        old_row[old_header_positions[header]]
                        if header in old_header_positions
                        and old_header_positions[header] < len(old_row)
                        else ""
                    )
                    for header in RESULT_HEADERS
                ]
            )

    for cell in worksheet[1]:
        cell.font = Font(bold=True)


def format_worksheet(worksheet) -> None:
    """设置结果工作表的阅读样式。"""

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    widths = (8, 14, 21, 14, 70, 48, 24, 16, 14, 16, 16, 12, 110)
    for index, width in enumerate(widths, start=1):
        worksheet.column_dimensions[chr(64 + index)].width = width


def write_result_excel(
    records: list[VideoRecord],
    excel_path: Path,
    keyword: str,
    result_type: str,
    time_range: str,
    crawled_at: str,
) -> tuple[int, int]:
    """按“视频名称 + 博主名称”去重，增量更新工作表。"""

    excel_path.parent.mkdir(parents=True, exist_ok=True)
    sheet_name = excel_sheet_name(keyword)
    if excel_path.exists():
        workbook = load_workbook(excel_path)
        worksheet = (
            workbook[sheet_name]
            if sheet_name in workbook.sheetnames
            else workbook.create_sheet(sheet_name)
        )
    else:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = sheet_name

    ensure_result_headers(worksheet)
    known_identities = worksheet_video_identities(worksheet)
    existing_numbers = [
        int(row[0])
        for row in worksheet.iter_rows(min_row=2, values_only=True)
        if row
        and isinstance(row[0], (int, float))
        and float(row[0]).is_integer()
    ]
    next_number = max(existing_numbers, default=0) + 1
    added_count = 0
    skipped_count = 0

    for record in records:
        identity = video_identity(record["title"], record["author_name"])
        if identity in known_identities:
            skipped_count += 1
            continue

        video_url = f"{DOUYIN_VIDEO_URL_PREFIX}{record['video_id']}"
        worksheet.append(
            [
                next_number,
                result_type,
                crawled_at,
                time_range,
                record["title"],
                video_url,
                record["author_name"],
                record["total_followers"],
                record["hotness"],
                record["new_views"],
                record["new_likes"],
                record["like_rate"],
                record.get("top_comments", ""),
            ]
        )
        url_cell = worksheet.cell(row=worksheet.max_row, column=6)
        url_cell.hyperlink = str(url_cell.value)
        url_cell.style = "Hyperlink"
        known_identities.add(identity)
        next_number += 1
        added_count += 1

    format_worksheet(worksheet)
    workbook.save(excel_path)
    return added_count, skipped_count
