"""调用视频提取接口，将口播文本增量写入结果 Excel。"""

import argparse
import json
import sys
import time
from collections.abc import Callable
from copy import copy
from pathlib import Path
from typing import Any
from urllib.error import HTTPError, URLError
from urllib.request import Request, urlopen

from openpyxl import load_workbook

from douhot_crawler.config import COOKIE_CONFIG_PATH, RESULT_EXCEL_PATH

DEFAULT_EXCEL_PATH = RESULT_EXCEL_PATH
DEFAULT_COOKIE_PATH = COOKIE_CONFIG_PATH
EXTRACT_API_URL = "http://examples:28600/api/v1/videos/extract"
VIDEO_URL_HEADER = "视频的url"
TRANSCRIPT_HEADER = "视频口播"


def parse_args() -> argparse.Namespace:
    """解析口播提取任务参数。"""

    parser = argparse.ArgumentParser(
        description="调用视频提取接口，将口播文本写入结果 Excel"
    )
    parser.add_argument(
        "--excel",
        type=Path,
        default=DEFAULT_EXCEL_PATH,
        help="结果 Excel 路径",
    )
    parser.add_argument(
        "--cookie-file",
        type=Path,
        default=DEFAULT_COOKIE_PATH,
        help="接口 cookie 文件路径（默认：cookie.config）",
    )
    parser.add_argument(
        "--sheet",
        action="append",
        help="仅处理指定 Sheet；可重复传入，默认处理所有 Sheet",
    )
    parser.add_argument(
        "--callback-url",
        default="",
        help="透传给接口的 callback_url（默认：空）",
    )
    parser.add_argument(
        "--timeout",
        type=float,
        default=90.0,
        help="单条接口请求超时秒数（默认：90）",
    )
    parser.add_argument(
        "--delay",
        type=float,
        default=0.0,
        help="每条请求后的等待秒数（默认：0）",
    )
    parser.add_argument(
        "--limit",
        type=int,
        help="本次最多处理的空口播记录数，用于分批执行",
    )
    parser.add_argument(
        "--overwrite",
        action="store_true",
        help="重新提取已有口播的记录",
    )
    return parser.parse_args()


def read_cookie(cookie_path: Path) -> str:
    """读取 cookie，不在日志或异常中暴露其内容。"""

    if not cookie_path.is_file():
        raise FileNotFoundError(f"没有找到 cookie 文件：{cookie_path}")

    cookie = cookie_path.read_text(encoding="utf-8").strip()
    if not cookie:
        raise ValueError(f"cookie 文件为空：{cookie_path}")
    return cookie


def extract_transcript(
    share_link: str,
    cookie: str,
    callback_url: str,
    timeout: float,
) -> str:
    """请求接口并返回视频口播文本。"""

    payload = json.dumps(
        {
            "share_link": share_link,
            "cookie": cookie,
            "callback_url": callback_url,
        },
        ensure_ascii=False,
    ).encode("utf-8")
    request = Request(
        EXTRACT_API_URL,
        data=payload,
        headers={"accept": "application/json", "Content-Type": "application/json"},
        method="POST",
    )

    try:
        with urlopen(request, timeout=timeout) as response:
            result: Any = json.load(response)
    except HTTPError as exc:
        raise RuntimeError(f"接口返回 HTTP {exc.code}") from exc
    except URLError as exc:
        raise RuntimeError(f"接口请求失败：{exc.reason}") from exc

    if not isinstance(result, dict):
        raise RuntimeError("接口返回不是 JSON 对象")
    if result.get("success") is not True:
        raise RuntimeError(f"接口未成功返回：{result.get('message', result)}")

    transcript = result.get("transcript")
    if not isinstance(transcript, str) or not transcript.strip():
        raise RuntimeError("接口成功但没有返回 transcript")
    return transcript.strip()


def transcript_column(worksheet) -> int:
    """获取或创建“视频口播”列，并返回列号。"""

    headers = [cell.value for cell in worksheet[1]]
    if TRANSCRIPT_HEADER in headers:
        return headers.index(TRANSCRIPT_HEADER) + 1

    column = worksheet.max_column + 1
    cell = worksheet.cell(row=1, column=column, value=TRANSCRIPT_HEADER)
    font = copy(cell.font)
    font.bold = True
    cell.font = font
    worksheet.column_dimensions[cell.column_letter].width = 110
    return column


def select_sheets(workbook, requested_sheets: list[str] | None):
    """确定本次处理的 Sheet，校验用户指定名称。"""

    if not requested_sheets:
        return list(workbook.worksheets)

    missing = sorted(set(requested_sheets) - set(workbook.sheetnames))
    if missing:
        raise ValueError(f"没有找到 Sheet：{', '.join(missing)}")
    return [workbook[name] for name in requested_sheets]


def analyze_excel(
    args: argparse.Namespace,
    *,
    stop_requested: Callable[[], bool] | None = None,
) -> tuple[int, int, int]:
    """提取并写入所有待处理视频口播，返回成功、跳过、失败数。"""

    if args.timeout <= 0:
        raise ValueError("--timeout 必须大于 0")
    if args.delay < 0:
        raise ValueError("--delay 不能小于 0")
    if args.limit is not None and args.limit <= 0:
        raise ValueError("--limit 必须大于 0")
    if not args.excel.is_file():
        raise FileNotFoundError(f"没有找到结果 Excel：{args.excel}")

    cookie = read_cookie(args.cookie_file)
    workbook = load_workbook(args.excel)
    success_count = 0
    skipped_count = 0
    failed_count = 0
    processed_count = 0

    try:
        for worksheet in select_sheets(workbook, args.sheet):
            headers = [cell.value for cell in worksheet[1]]
            if VIDEO_URL_HEADER not in headers:
                print(f"跳过 Sheet {worksheet.title}：没有“{VIDEO_URL_HEADER}”列")
                continue

            url_column = headers.index(VIDEO_URL_HEADER) + 1
            had_transcript_column = TRANSCRIPT_HEADER in headers
            text_column = transcript_column(worksheet)
            if not had_transcript_column:
                workbook.save(args.excel)
            print(f"\n开始处理 Sheet：{worksheet.title}")

            for row_number in range(2, worksheet.max_row + 1):
                if stop_requested and stop_requested():
                    print("已收到安全停止请求，已完成的口播已写入 Excel")
                    return success_count, skipped_count, failed_count
                if args.limit is not None and processed_count >= args.limit:
                    break

                share_link = worksheet.cell(row=row_number, column=url_column).value
                existing_text = worksheet.cell(row=row_number, column=text_column).value
                if not isinstance(share_link, str) or not share_link.strip():
                    skipped_count += 1
                    continue
                if existing_text and not args.overwrite:
                    skipped_count += 1
                    continue

                processed_count += 1
                try:
                    transcript = extract_transcript(
                        share_link=share_link.strip(),
                        cookie=cookie,
                        callback_url=args.callback_url,
                        timeout=args.timeout,
                    )
                    cell = worksheet.cell(row=row_number, column=text_column)
                    cell.value = transcript
                    alignment = copy(cell.alignment)
                    alignment.wrap_text = True
                    alignment.vertical = "top"
                    cell.alignment = alignment
                    workbook.save(args.excel)
                    success_count += 1
                    print(f"  第 {row_number} 行口播已写入")
                except Exception as exc:
                    failed_count += 1
                    print(f"  第 {row_number} 行提取失败：{exc}", file=sys.stderr)

                if args.delay:
                    time.sleep(args.delay)

            if args.limit is not None and processed_count >= args.limit:
                break
    finally:
        workbook.close()

    return success_count, skipped_count, failed_count


def main() -> None:
    args = parse_args()
    success_count, skipped_count, failed_count = analyze_excel(args)
    print(
        f"\n完成：口播写入 {success_count} 条，"
        f"跳过 {skipped_count} 条，失败 {failed_count} 条"
    )


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        raise SystemExit(130)
    except Exception as exc:
        print(f"\n错误：{exc}", file=sys.stderr)
        raise SystemExit(1)
