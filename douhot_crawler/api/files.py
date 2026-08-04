from __future__ import annotations

import hashlib
import shutil
from contextlib import contextmanager
from pathlib import Path
from typing import Iterator

from filelock import FileLock, Timeout
from openpyxl import load_workbook

from .errors import ApiError


@contextmanager
def workbook_lock(path: Path) -> Iterator[None]:
    """Acquire a non-waiting cross-process lock for one workbook."""

    path.parent.mkdir(parents=True, exist_ok=True)
    lock = FileLock(f"{path}.lock")
    try:
        lock.acquire(timeout=0)
    except Timeout as exc:
        raise ApiError(
            "WORKBOOK_BUSY",
            "目标 Excel 正在被其他任务写入",
            status_code=409,
        ) from exc
    try:
        yield
    finally:
        lock.release()


def workbook_metadata(path: Path) -> dict[str, object]:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        row_count = sum(max(sheet.max_row - 1, 0) for sheet in workbook.worksheets)
    finally:
        workbook.close()
    return {
        "path": str(path),
        "row_count": row_count,
        "sha256": digest.hexdigest(),
    }


def safe_remove_task_directory(path: Path, tasks_root: Path) -> None:
    resolved = path.resolve()
    allowed = tasks_root.resolve()
    if resolved != allowed and resolved.is_relative_to(allowed):
        shutil.rmtree(resolved, ignore_errors=True)
