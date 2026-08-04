from __future__ import annotations

import argparse
import asyncio
import hashlib
import json
import re
from contextlib import suppress
from datetime import datetime, timedelta
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path
from typing import Any
from zoneinfo import ZoneInfo

from openpyxl import load_workbook

from douhot_crawler.browser.setup import chromium_status
from douhot_crawler.core.models import RunOptions
from douhot_crawler.core.storage import excel_sheet_name
from douhot_crawler.crawling.runner import run as run_crawler
from douhot_crawler.transcript.analyzer import analyze_excel

from .clients import ExternalApiClient
from .config import ApiSettings
from .errors import ApiError, TaskPaused
from .files import workbook_lock, workbook_metadata, safe_remove_task_directory
from .models import (
    AnalyzeTaskRequest,
    CrawlTaskRequest,
    PipelineTaskRequest,
    TaskKind,
    TaskStatus,
    UploadTaskRequest,
)
from .store import ApiTaskStore


UPLOAD_HEADERS = {
    "videoName": "视频名称",
    "videoUrl": "视频的url",
    "authorName": "博主名称",
    "followerCount": "总粉丝数",
    "heatValue": "热度值",
    "newPlayCount": "新增播放量",
    "newLikeCount": "新增点赞量",
    "likeRate": "点赞率",
    "highPraiseComment": "高赞评论",
    "videoOral": "视频口播",
}
SHANGHAI_TIMEZONE = ZoneInfo("Asia/Shanghai")


def parse_follower_count(value: object) -> tuple[int, bool]:
    """Parse Chinese display units into the integer required downstream."""

    if value is None or isinstance(value, bool):
        return 0, True
    text = str(value).strip().replace(",", "").replace("，", "")
    if not text:
        return 0, True
    match = re.fullmatch(r"([+-]?\d+(?:\.\d+)?)\s*([千千万亿]?)\s*(?:粉丝)?", text)
    if not match:
        return 0, True
    factor = {"": 1, "千": 1_000, "万": 10_000, "亿": 100_000_000}[match.group(2)]
    try:
        number = Decimal(match.group(1)) * factor
    except InvalidOperation:
        return 0, True
    if number < 0:
        return 0, True
    return int(number.quantize(Decimal("1"), rounding=ROUND_HALF_UP)), False


class ApiTaskService:
    """Own the durable FIFO worker and all API task orchestration."""

    def __init__(
        self,
        settings: ApiSettings,
        *,
        store: ApiTaskStore | None = None,
        client: ExternalApiClient | None = None,
    ) -> None:
        self.settings = settings
        self.data_root = settings.data_root
        self.tasks_root = self.data_root / "tasks"
        self.tasks_root.mkdir(parents=True, exist_ok=True)
        self.store = store or ApiTaskStore(self.data_root / "tasks.sqlite3")
        self.client = client or ExternalApiClient(settings)
        self._wake = asyncio.Event()
        self._closing = False
        self._worker: asyncio.Task[None] | None = None
        self._scheduler: asyncio.Task[None] | None = None
        self._next_daily_run: datetime | None = None
        self.browser_ok, self.browser_message = chromium_status()

    async def start(self) -> None:
        self.cleanup()
        if self._worker is None:
            self._worker = asyncio.create_task(self._worker_loop(), name="douhot-api-worker")
            self._wake.set()
        if self.settings.daily_enabled and self._scheduler is None:
            self._scheduler = asyncio.create_task(
                self._scheduler_loop(), name="douhot-api-daily-scheduler"
            )

    async def close(self) -> None:
        self._closing = True
        if self._scheduler:
            self._scheduler.cancel()
            with suppress(asyncio.CancelledError):
                await self._scheduler
        self.store.request_worker_shutdown()
        self._wake.set()
        if self._worker:
            await self._worker
        await self.client.close()

    def next_daily_run(self, now: datetime | None = None) -> datetime:
        current = now or datetime.now(SHANGHAI_TIMEZONE)
        if current.tzinfo is None:
            current = current.replace(tzinfo=SHANGHAI_TIMEZONE)
        else:
            current = current.astimezone(SHANGHAI_TIMEZONE)
        hour, minute = self.settings.daily_hour_minute
        target = current.replace(hour=hour, minute=minute, second=0, microsecond=0)
        if target <= current:
            target += timedelta(days=1)
        return target

    async def _scheduler_loop(self) -> None:
        while not self._closing:
            target = self.next_daily_run()
            self._next_daily_run = target
            delay = max((target - datetime.now(SHANGHAI_TIMEZONE)).total_seconds(), 0)
            await asyncio.sleep(delay)
            if self._closing:
                return
            task, created = self.create_pipeline(PipelineTaskRequest())
            if created:
                self._log(task["task_id"], "上海时区每日定时任务已触发")
            else:
                self._log(task["task_id"], "已有活动流水线，每日定时任务未重复创建")

    def _task_dir(self, task_id: str) -> Path:
        return self.tasks_root / task_id

    def _workbook(self, task_id: str) -> Path:
        return self._task_dir(task_id) / "result.xlsx"

    def _log(self, task_id: str, message: str) -> None:
        directory = self._task_dir(task_id)
        directory.mkdir(parents=True, exist_ok=True)
        timestamp = datetime.now().astimezone().isoformat(timespec="seconds")
        with (directory / "task.log").open("a", encoding="utf-8") as handle:
            handle.write(f"{timestamp} {message}\n")

    def cleanup(self) -> None:
        for task_id in self.store.old_artifact_task_ids(
            self.settings.artifact_retention_days
        ):
            safe_remove_task_directory(self._task_dir(task_id), self.tasks_root)
        self.store.cleanup_metadata(self.settings.metadata_retention_days)

    def status(self, task_id: str) -> dict[str, Any]:
        return self.store.get(task_id)

    async def keywords(self) -> list[str]:
        return await self.client.fetch_keywords()

    def create_crawl(self, request: CrawlTaskRequest) -> dict[str, Any]:
        task = self.store.create(TaskKind.CRAWL, request.model_dump(mode="json"))
        self._log(task["task_id"], "爬取任务已进入队列")
        self._wake.set()
        return task

    def create_analyze(self, request: AnalyzeTaskRequest) -> dict[str, Any]:
        source = self.store.get(request.crawl_task_id)
        if source["kind"] != TaskKind.CRAWL or source["status"] not in {
            TaskStatus.SUCCEEDED,
            TaskStatus.SUCCEEDED_WITH_WARNINGS,
        }:
            raise ApiError(
                "CRAWL_NOT_READY", "crawl_task_id 对应的爬取任务尚未成功", status_code=409
            )
        workbook = self._workbook(request.crawl_task_id)
        if not workbook.is_file():
            raise ApiError("ARTIFACT_NOT_FOUND", "爬取结果 Excel 不存在", status_code=404)
        with workbook_lock(workbook):
            pass
        task = self.store.create(TaskKind.ANALYZE, request.model_dump(mode="json"))
        self._log(task["task_id"], "口播任务已进入队列")
        self._wake.set()
        return task

    def create_pipeline(self, request: PipelineTaskRequest) -> tuple[dict[str, Any], bool]:
        existing = self.store.active_pipeline()
        if existing:
            return existing, False
        task = self.store.create(TaskKind.PIPELINE, request.model_dump(mode="json"))
        self._log(task["task_id"], "流水线任务已进入队列")
        self._wake.set()
        return task, True

    def _source_workbook(self, source_task_id: str) -> Path:
        source = self.store.get(source_task_id)
        if source["status"] in {
            TaskStatus.QUEUED,
            TaskStatus.RUNNING,
            TaskStatus.PAUSING,
        }:
            raise ApiError(
                "SOURCE_TASK_BUSY",
                "源任务仍在排队或写入 Excel，请等待其完成或暂停",
                status_code=409,
            )
        kind = TaskKind(source["kind"])
        if kind in {TaskKind.CRAWL, TaskKind.PIPELINE}:
            owner_task_id = source_task_id
        elif kind == TaskKind.ANALYZE:
            owner_task_id = str(source["params"].get("crawl_task_id") or "")
        else:
            raise ApiError(
                "INVALID_SOURCE_TASK",
                "source_task_id 必须是 crawl、analyze 或 pipeline 任务",
                status_code=400,
            )
        workbook = self._workbook(owner_task_id)
        if not workbook.is_file():
            raise ApiError("ARTIFACT_NOT_FOUND", "源任务的 Excel 不存在", status_code=404)
        with workbook_lock(workbook):
            pass
        return workbook

    def create_upload(self, request: UploadTaskRequest) -> dict[str, Any]:
        self._source_workbook(request.source_task_id)
        task = self.store.create(TaskKind.UPLOAD, request.model_dump(mode="json"))
        self._log(task["task_id"], "Excel 全量发送任务已进入队列")
        self._wake.set()
        return task

    def pause(self, task_id: str) -> dict[str, Any]:
        task = self.store.request_pause(task_id)
        self._log(task_id, "收到用户暂停请求")
        return task

    def resume(self, task_id: str) -> dict[str, Any]:
        task = self.store.resume(task_id)
        self._log(task_id, "任务已重新进入队列")
        self._wake.set()
        return task

    async def _worker_loop(self) -> None:
        while True:
            if self._closing:
                return
            task = self.store.claim_next()
            if task is None:
                if self._closing:
                    return
                self._wake.clear()
                try:
                    await asyncio.wait_for(self._wake.wait(), timeout=2.0)
                except TimeoutError:
                    pass
                continue
            await self._execute(task)

    async def _execute(self, task: dict[str, Any]) -> None:
        task_id = task["task_id"]
        self._log(task_id, f"开始执行 {task['kind']} 任务")
        try:
            kind = TaskKind(task["kind"])
            if kind == TaskKind.CRAWL:
                await self._execute_crawl(task_id, CrawlTaskRequest.model_validate(task["params"]))
            elif kind == TaskKind.ANALYZE:
                await self._execute_analyze(
                    task_id, AnalyzeTaskRequest.model_validate(task["params"])
                )
            elif kind == TaskKind.UPLOAD:
                await self._execute_upload(
                    task_id, UploadTaskRequest.model_validate(task["params"])
                )
            else:
                await self._execute_pipeline(
                    task_id, PipelineTaskRequest.model_validate(task["params"])
                )
        except TaskPaused as exc:
            if exc.reason == "shutdown":
                if task["kind"] == TaskKind.PIPELINE:
                    self.store.update(
                        task_id,
                        status=TaskStatus.QUEUED,
                        pause_requested=False,
                        pause_reason="shutdown",
                    )
                else:
                    self.store.fail(task_id, "服务关闭，手动任务已终止")
            else:
                self.store.mark_paused(task_id, exc.reason)
            self._log(task_id, f"任务已停止：{exc.reason}")
        except Exception as exc:
            self.store.fail(task_id, str(exc))
            self._log(task_id, f"任务失败：{exc}")

    def _raise_if_paused(self, task_id: str, reason: str = "user") -> None:
        task = self.store.get(task_id)
        if task["pause_requested"]:
            raise TaskPaused(task.get("pause_reason") or reason)

    def _artifact(self, workbook: Path) -> dict[str, Any]:
        metadata = workbook_metadata(workbook)
        metadata["path"] = str(workbook.relative_to(self.data_root))
        return metadata

    async def _crawl_keyword(
        self,
        task_id: str,
        request: CrawlTaskRequest,
        workbook: Path,
    ) -> dict[str, Any]:
        self._raise_if_paused(task_id)
        target = request.limit or self.settings.max_videos_per_keyword
        existing = 0
        if workbook.is_file():
            with workbook_lock(workbook):
                book = load_workbook(workbook, read_only=True, data_only=True)
                try:
                    sheet_name = excel_sheet_name(request.keyword)
                    if sheet_name in book.sheetnames:
                        existing = max(book[sheet_name].max_row - 1, 0)
                finally:
                    book.close()
        remaining = max(target - existing, 0)
        if remaining == 0:
            return {
                "excel_path": str(workbook),
                "added_count": 0,
                "skipped_count": existing,
                "sheet": excel_sheet_name(request.keyword),
                "stopped": False,
            }
        cookie = await self.client.fetch_cookie(0)
        options = RunOptions(
            keyword=request.keyword,
            input_timeout=request.input_timeout,
            detail_delay=request.detail_delay,
            result_type=request.result_type,
            time_range=request.time_range,
            headless=True,
        )

        async def progress(values: dict[str, Any]) -> None:
            self.store.update_progress(task_id, phase="crawl", **values)

        with workbook_lock(workbook):
            result = await run_crawler(
                options,
                excel_path=workbook,
                max_results=remaining,
                browser_cookie=cookie,
                stop_requested=lambda: self.store.should_pause(task_id),
                progress_callback=progress,
            )
        if result.get("stopped") or self.store.should_pause(task_id):
            state = self.store.get(task_id)
            raise TaskPaused(state.get("pause_reason") or "user")
        return result

    async def _execute_crawl(self, task_id: str, request: CrawlTaskRequest) -> None:
        workbook = self._workbook(task_id)
        result = await self._crawl_keyword(task_id, request, workbook)
        artifact = self._artifact(workbook)
        result["artifact"] = artifact
        self.store.finish(task_id, result=result, artifact_path=workbook)
        self._log(task_id, f"爬取完成，新增 {result['added_count']} 条")

    def _analyze_args(
        self,
        request: AnalyzeTaskRequest,
        workbook: Path,
        cookie: str,
        task_id: str,
    ) -> argparse.Namespace:
        def progress(values: dict[str, Any]) -> None:
            self.store.update_progress(task_id, phase="analyze", **values)

        return argparse.Namespace(
            excel=workbook,
            cookie=cookie,
            cookie_file=None,
            sheet=request.sheets,
            callback_url="",
            timeout=request.timeout,
            delay=request.delay,
            limit=request.limit,
            overwrite=request.overwrite,
            api_url=self.settings.external_url("extract_api_url"),
            progress_callback=progress,
        )

    async def _analyze_workbook(
        self,
        task_id: str,
        request: AnalyzeTaskRequest,
        workbook: Path,
    ) -> tuple[int, int, int]:
        self._raise_if_paused(task_id)
        cookie = await self.client.fetch_cookie(1)
        args = self._analyze_args(request, workbook, cookie, task_id)
        with workbook_lock(workbook):
            counts = await asyncio.to_thread(
                analyze_excel,
                args,
                stop_requested=lambda: self.store.should_pause(task_id),
            )
        if self.store.should_pause(task_id):
            state = self.store.get(task_id)
            raise TaskPaused(state.get("pause_reason") or "user")
        return counts

    async def _execute_analyze(
        self, task_id: str, request: AnalyzeTaskRequest
    ) -> None:
        workbook = self._workbook(request.crawl_task_id)
        succeeded, skipped, failed = await self._analyze_workbook(task_id, request, workbook)
        if failed:
            self.store.add_warning(task_id, f"有 {failed} 条视频口播提取失败")
        result = {
            "succeeded": succeeded,
            "skipped": skipped,
            "failed": failed,
            "artifact": self._artifact(workbook),
        }
        self.store.finish(task_id, result=result, artifact_path=workbook)
        self._log(task_id, f"口播完成：成功 {succeeded}，跳过 {skipped}，失败 {failed}")

    async def _execute_upload(
        self, task_id: str, request: UploadTaskRequest
    ) -> None:
        workbook = self._source_workbook(request.source_task_id)
        with workbook_lock(workbook):
            book = load_workbook(workbook, read_only=True, data_only=True)
            try:
                if request.sheets:
                    missing = sorted(set(request.sheets) - set(book.sheetnames))
                    if missing:
                        raise ValueError(f"Excel 中不存在 Sheet：{', '.join(missing)}")
                    sheet_names = request.sheets
                else:
                    sheet_names = list(book.sheetnames)
            finally:
                book.close()

        eligible = 0
        sent = 0
        for index, sheet_name in enumerate(sheet_names, start=1):
            self._raise_if_paused(task_id)
            self.store.update_progress(
                task_id,
                phase="upload",
                current=index,
                total=len(sheet_names),
                sheet=sheet_name,
            )
            counts = await self._upload_sheet(task_id, workbook, sheet_name)
            eligible += counts["eligible"]
            sent += counts["sent"]

        result = {
            "sheets": sheet_names,
            "eligible_rows": eligible,
            "sent_rows": sent,
            "artifact": self._artifact(workbook),
        }
        self.store.finish(task_id, result=result, artifact_path=workbook)
        self._log(task_id, f"Excel 发送完成：合格 {eligible} 条，已发送 {sent} 条")

    async def _execute_pipeline(
        self, task_id: str, request: PipelineTaskRequest
    ) -> None:
        checkpoints = self.store.pipeline_keywords(task_id)
        if not checkpoints:
            keywords = request.keywords or await self.client.fetch_keywords()
            if not keywords:
                raise RuntimeError("热点关键词接口没有返回有效关键词")
            self.store.set_pipeline_keywords(task_id, keywords)
            checkpoints = self.store.pipeline_keywords(task_id)
            self.store.update_progress(task_id, phase="keywords", total=len(keywords), current=0)

        workbook = self._workbook(task_id)
        successful_keywords = 0
        for checkpoint in checkpoints:
            position = checkpoint["position"]
            keyword = checkpoint["keyword"]
            self._raise_if_paused(task_id)
            self.store.update_progress(
                task_id,
                phase="keyword",
                current=position + 1,
                total=len(checkpoints),
                keyword=keyword,
            )
            try:
                if not checkpoint["crawl_done"]:
                    crawl_request = CrawlTaskRequest(
                        keyword=keyword,
                        result_type=request.result_type,
                        time_range=request.time_range,
                        input_timeout=request.input_timeout,
                        detail_delay=request.detail_delay,
                        limit=request.limit_per_keyword,
                    )
                    await self._crawl_keyword(task_id, crawl_request, workbook)
                    self.store.update_pipeline_keyword(task_id, position, crawl_done=True)

                if not checkpoint["analyze_done"]:
                    analyze_request = AnalyzeTaskRequest(
                        crawl_task_id=task_id,
                        sheets=[excel_sheet_name(keyword)],
                        overwrite=request.overwrite_transcript,
                    )
                    succeeded, skipped, failed = await self._analyze_workbook(
                        task_id, analyze_request, workbook
                    )
                    if failed:
                        self.store.add_warning(
                            task_id, f"关键词 {keyword} 有 {failed} 条口播提取失败"
                        )
                    self.store.update_pipeline_keyword(task_id, position, analyze_done=True)
                    self.store.update_progress(
                        task_id,
                        phase="analyze",
                        keyword=keyword,
                        transcript_succeeded=succeeded,
                        transcript_skipped=skipped,
                        transcript_failed=failed,
                    )

                if not checkpoint["upload_done"]:
                    await self._upload_sheet(task_id, workbook, excel_sheet_name(keyword))
                    self.store.update_pipeline_keyword(task_id, position, upload_done=True)
                successful_keywords += 1
                self._log(task_id, f"关键词完成：{keyword}")
            except TaskPaused:
                raise
            except Exception as exc:
                self.store.update_pipeline_keyword(
                    task_id, position, last_error=str(exc)
                )
                self.store.add_warning(task_id, f"关键词 {keyword} 处理失败：{exc}")
                self._log(task_id, f"关键词失败，继续下一个：{keyword}：{exc}")

        if successful_keywords == 0:
            raise RuntimeError("全部关键词处理失败")
        result = {
            "keywords_total": len(checkpoints),
            "keywords_succeeded": successful_keywords,
            "keywords_failed": len(checkpoints) - successful_keywords,
            "artifact": self._artifact(workbook),
        }
        self.store.finish(task_id, result=result, artifact_path=workbook)
        self._log(task_id, "流水线任务完成")

    def _sheet_payloads(
        self, task_id: str, workbook: Path, sheet_name: str
    ) -> list[tuple[dict[str, Any], str]]:
        with workbook_lock(workbook):
            book = load_workbook(workbook, read_only=True, data_only=True)
            try:
                if sheet_name not in book.sheetnames:
                    raise RuntimeError(f"Excel 中不存在 Sheet：{sheet_name}")
                sheet = book[sheet_name]
                headers = [str(cell.value or "").strip() for cell in sheet[1]]
                positions = {
                    name: headers.index(header)
                    for name, header in UPLOAD_HEADERS.items()
                    if header in headers
                }
                required_headers = set(UPLOAD_HEADERS) - set(positions)
                if required_headers:
                    raise RuntimeError(f"Excel 缺少发送字段：{', '.join(sorted(required_headers))}")
                output: list[tuple[dict[str, Any], str]] = []
                for row_number, row in enumerate(
                    sheet.iter_rows(min_row=2, values_only=True), start=2
                ):
                    values = {
                        name: row[index] if index < len(row) else None
                        for name, index in positions.items()
                    }
                    if any(
                        not str(values[name] or "").strip()
                        for name in ("videoName", "videoUrl", "authorName", "videoOral")
                    ):
                        continue
                    follower_count, warned = parse_follower_count(values["followerCount"])
                    if warned:
                        self.store.add_warning(
                            task_id,
                            f"{sheet_name} 第 {row_number} 行粉丝数无法解析，按 0 发送",
                        )
                    payload: dict[str, Any] = {"type": 0, "keyword": sheet_name}
                    for name in UPLOAD_HEADERS:
                        if name == "followerCount":
                            payload[name] = follower_count
                        else:
                            payload[name] = str(values[name] or "").strip()
                    serialized = json.dumps(
                        payload, ensure_ascii=False, sort_keys=True, separators=(",", ":")
                    )
                    output.append(
                        (payload, hashlib.sha256(serialized.encode("utf-8")).hexdigest())
                    )
                return output
            finally:
                book.close()

    async def _upload_sheet(
        self, task_id: str, workbook: Path, sheet_name: str
    ) -> dict[str, int]:
        payloads = self._sheet_payloads(task_id, workbook, sheet_name)
        pending = [
            (payload, digest)
            for payload, digest in payloads
            if not self.store.delivery_sent(
                task_id, sheet_name, payload["videoUrl"], digest
            )
        ]
        already_sent = len(payloads) - len(pending)
        sent = already_sent
        for offset in range(0, len(pending), self.settings.upload_batch_size):
            self._raise_if_paused(task_id)
            batch = pending[offset : offset + self.settings.upload_batch_size]
            try:
                await self.client.upload_rankings([payload for payload, _ in batch])
            except Exception as exc:
                self.store.mark_deliveries_failed(
                    task_id,
                    [
                        (sheet_name, payload["videoUrl"], digest)
                        for payload, digest in batch
                    ],
                    str(exc),
                )
                self.store.update_progress(
                    task_id,
                    phase="upload",
                    keyword=sheet_name,
                    sent=sent,
                    total=len(payloads),
                )
                raise TaskPaused("upload_failure") from exc
            self.store.mark_deliveries_sent(
                task_id,
                [
                    (sheet_name, payload["videoUrl"], digest)
                    for payload, digest in batch
                ],
            )
            sent += len(batch)
            self.store.update_progress(
                task_id,
                phase="upload",
                keyword=sheet_name,
                sent=sent,
                total=len(payloads),
            )
        return {"eligible": len(payloads), "sent": sent}

    def health(self) -> dict[str, Any]:
        return {
            "status": "ok" if self.store.ping() and self.browser_ok else "degraded",
            "worker_running": bool(self._worker and not self._worker.done()),
            "database_ok": self.store.ping(),
            "browser_ok": self.browser_ok,
            "external_urls_configured": True,
            "scheduler_overlap": self.store.active_pipeline() is not None,
            "scheduler_enabled": self.settings.daily_enabled,
            "scheduler_time": self.settings.daily_time,
            "scheduler_timezone": "Asia/Shanghai",
            "scheduler_next_run_at": (
                self._next_daily_run.isoformat() if self._next_daily_run else None
            ),
        }
